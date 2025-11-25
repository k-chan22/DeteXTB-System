import streamlit as st
import folium
import pandas as pd
import streamlit.components.v1 as components
import io
import numpy as np
from Supabase import supabase
from folium import Element
from folium.plugins import HeatMap
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment
from collections import defaultdict
from dateutil.parser import parse


# Define the page as a function to be used by sidebar.py
def Heatmap(is_light=True):

    if "light_mode" not in st.session_state:
        st.session_state["light_mode"] = True  # Default theme

    if is_light is None:
        is_light = st.session_state["light_mode"] 

    if "heatmap_filters" not in st.session_state:
        st.session_state.heatmap_filters = {
            "date": datetime.today().date(),
            "barangay": "All",
            "age_group": "All",
            "sex": "All"
        }

    if "heatmap_date_filter" not in st.session_state:
        st.session_state["heatmap_date_filter"] = datetime.today().date()

    if "heatmap_barangay_filter" not in st.session_state:
        st.session_state["heatmap_barangay_filter"] = "All"

    if "heatmap_age_filter" not in st.session_state:
        st.session_state["heatmap_age_filter"] = 0

    if "heatmap_sex_filter" not in st.session_state:
        st.session_state["heatmap_sex_filter"] = "All"

    if "reset_triggered" not in st.session_state:
        st.session_state["reset_triggered"] = False

    if st.session_state.reset_triggered:
        st.session_state["heatmap_date_filter"] = datetime.today().date()
        st.session_state["heatmap_barangay_filter"] = "All"
        st.session_state["heatmap_age_filter"] = 0
        st.session_state["heatmap_sex_filter"] = "All"
        st.session_state["reset_triggered"] = False  # Clear flag
        st.rerun()

    heatmap_date_filter = st.session_state["heatmap_date_filter"]
    heatmap_barangay_filter = st.session_state["heatmap_barangay_filter"]
    heatmap_age_filter = st.session_state["heatmap_age_filter"]
    heatmap_sex_filter = st.session_state["heatmap_sex_filter"]


    # --- Constants Initialization ---

    # Incidence Rate -> Incidence Rate = (Number of New Cases / Population) * Multiplier
    incidence_rate = 0.00539

    # Mandaue City Population Data (2025 Census) = 388,002
    mandaue_barangay_population_2025 = {
        "Alang-Alang": 12251, "Bakilid": 4671, "Banilad": 19594, "Basak": 12554,
        "Cabancalan": 15818, "Cambaro": 9581, "Canduman": 24995, "Casili": 5756,
        "Casuntingan": 17956, "Centro (Poblacion)": 3171, "Cubacub": 14738, "Guizo": 7728,
        "Ibabao-Estancia": 7449, "Jagobiao": 12937, "Labogon": 21811, "Looc": 18536,
        "Maguikay": 15931, "Mantuyong": 5846, "Opao": 12798, "Pagsabungan": 21603,
        "Pakna-an": 32540, "Subangdaku": 18219, "Tabok": 20767, "Tawason": 7440,
        "Tingub": 6479, "Tipolo": 16822, "Umapad": 20011
    }


    # --- Helper functions ---

    # Function to generate heatmap summary report in Excel format
    def generate_heatmap_excel(data, risk_levels, selected_month, selected_year):
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Header Info
        ws.merge_cells('A1:J1')
        ws['A1'] = "Mandaue City Health Office"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='left')

        ws.merge_cells('A2:J2')
        ws['A2'] = "S.B. Cabahug, Mandaue City, Philippines."
        ws['A2'].alignment = Alignment(horizontal='left')

        ws.merge_cells('A3:J3')
        ws['A3'] = "Call us on: +63 (032) 230 4500 | FB: Mandaue City Public Affairs Office | Email: cmo@mandauecity.gov.ph"
        ws['A3'].alignment = Alignment(horizontal='left')

        ws.merge_cells('A5:J5')
        ws['A5'] = "DeteXTB: AI-Assisted Presumptive Tuberculosis Detection and Mapping System"
        ws['A5'].font = Font(bold=True, size=12)
        ws['A5'].alignment = Alignment(horizontal='left')

        # Reporting Period
        if selected_month == "All":
            reporting_period = f"Year {selected_year}"
        else:
            month_name = datetime(2000, selected_month, 1).strftime('%B')
            reporting_period = f"{month_name} {selected_year}"
        
        ws.merge_cells('A7:J7')
        ws['A7'] = f"Reporting Period: {reporting_period}"
        ws['A7'].font = Font(italic=True, size=11)
        ws['A7'].alignment = Alignment(horizontal='left')

        # Total Record with pastel red fill for emphasis
        ws['A9'] = "Total Records"
        ws['A9'].fill = PatternFill(start_color="F8B6B8", end_color="F8B6B8", fill_type="solid")
        ws['B9'] = len(data)
        ws['B9'].fill = PatternFill(start_color="F8B6B8", end_color="F8B6B8", fill_type="solid")

        # Prepare/Initialize counts
        barangay_data = defaultdict(lambda: {"count": 0})
        sex_counts = {"Male": 0, "Female": 0}
        age_groups = {
            "Children (0-14)": 0,
            "Youth/Young Adults (15-24)": 0,
            "Adults (25-64)": 0,
            "Elderly (65+)": 0,
            "Unknown": 0
        }

        for record in data:
            # Barangay
            brgy = record.get("MAP_BRGY", "Unknown")
            barangay_data[brgy]["count"] += 1

            # Sex
            sex = record.get("MAP_SEX", "Unknown")
            if sex not in sex_counts:
                sex_counts[sex] = 0
            sex_counts[sex] += 1

            # Age group
            age_group_str = record.get("MAP_AGE_GROUP", "Unknown")
            if age_group_str == "0-14":
                age_groups["Children (0-14)"] += 1
            elif age_group_str == "15-24":
                age_groups["Youth/Young Adults (15-24)"] += 1
            elif age_group_str == "25-64":
                age_groups["Adults (25-64)"] += 1
            elif age_group_str == "65+":
                age_groups["Elderly (65+)"] += 1
            else:
                age_groups["Unknown"] += 1

        # Table Headers
        ws['A11'] = "Barangay"
        ws['B11'] = "Count"
        ws['C11'] = "Risk Level"
        ws['E11'] = "Sex"
        ws['F11'] = "Count"
        ws['H11'] = "Age Group"
        ws['I11'] = "Count"

        header_cells = ['A11','B11','C11','E11','F11','H11','I11']
        for cell in header_cells:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = PatternFill(start_color="FECEAB", end_color="FECEAB", fill_type="solid")

        # Risk level font colors
        risk_text_colors = {
            "Low": "008000",      # Green
            "Moderate": "FF8C00", # Orange
            "High": "DC143C"      # Red
        }

        # Fill data rows
        max_len = max(len(barangay_data), len(sex_counts), len(age_groups))
        barangay_list = list(barangay_data.items())
        sex_list = list(sex_counts.items())
        age_list = list(age_groups.items())

        for i in range(max_len):
            row = i + 12
            if i < len(barangay_list):
                brgy, info = barangay_list[i]
                ws[f'A{row}'] = brgy
                ws[f'B{row}'] = info["count"]
                ws[f'B{row}'].alignment = Alignment(horizontal="left")

                risk_label, risk_color_name = risk_levels.get(brgy, ("Low Risk", "green"))
                risk_level = risk_label.replace(" Risk", "") if " Risk" in risk_label else risk_label
                ws[f'C{row}'] = risk_level
                if risk_level in risk_text_colors:
                    ws[f'C{row}'].font = Font(color=risk_text_colors[risk_level], bold=True)

            if i < len(sex_list):
                sex, count = sex_list[i]
                ws[f'E{row}'] = sex
                ws[f'F{row}'] = count
                ws[f'F{row}'].alignment = Alignment(horizontal="left")

            if i < len(age_list):
                age_group, count = age_list[i]
                ws[f'H{row}'] = age_group
                ws[f'I{row}'] = count
                ws[f'I{row}'].alignment = Alignment(horizontal="left")

        # Disclaimer section
        disclaimer_row_1 = max_len + 14
        disclaimer_row_2 = disclaimer_row_1 + 1

        ws.merge_cells(f"A{disclaimer_row_1}:J{disclaimer_row_1}")
        ws[f"A{disclaimer_row_1}"] = (
            "The heatmap and targets are based on estimated TB incidence (0.539%)."
        )
        ws[f"A{disclaimer_row_1}"].font = Font(italic=True, color="808080")
        ws[f"A{disclaimer_row_1}"].alignment = Alignment(horizontal='left')

        ws.merge_cells(f"A{disclaimer_row_2}:J{disclaimer_row_2}")
        ws[f"A{disclaimer_row_2}"] = (
            "These may not be real-time counts and should be verified with local surveillance before operational use."
        )
        ws[f"A{disclaimer_row_2}"].font = Font(italic=True, color="808080")
        ws[f"A{disclaimer_row_2}"].alignment = Alignment(horizontal='left')

        # Timestamp
        timestamp_row = disclaimer_row_2 + 3
        ws.merge_cells(f'A{timestamp_row}:I{timestamp_row}')
        ws[f'A{timestamp_row}'] = f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{timestamp_row}'].font = Font(italic=True)
        ws[f'A{timestamp_row}'].alignment = Alignment(horizontal='left')

        # Column Widths
        column_widths = {
            "A": 20, "B": 10, "C": 12, "D": 2,
            "E": 10, "F": 10, "G": 2,
            "H": 28, "I": 10
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


# ---------- Main Content ----------

    # Theme Toggle UI
    col_title, col_toggle = st.columns([6, 1])
    with col_title:
        st.markdown("<h4>Heatmap</h4>", unsafe_allow_html=True)
    with col_toggle:
        new_toggle = st.toggle("🌙", value=is_light, key="theme_toggle", label_visibility="collapsed")

    if new_toggle != st.session_state["light_mode"]:
        st.session_state["light_mode"] = new_toggle
        st.rerun()  # <-- Rerun to apply theme across app

    is_light = st.session_state["light_mode"]

    # Theme-based color variables
    bg_color = "white" if is_light else "#0e0e0e"
    text_color = "black" if is_light else "white"
    input_bg = "white" if is_light else "#1e1e1e"
    input_border = "black" if is_light else "white"
    placeholder_color = "#b0b0b0" if is_light else "#888"
    calendar_dropdown_bg = "white" if is_light else "#222"
    calendar_border = "black" if is_light else "none"
    select_bg = "white" if is_light else "#1e1e1e"
    select_hover = "#f0f0f0" if is_light else "#2a2a2a"
    error_text = "#b71c1c" if is_light else "#ff8a80"
    button_bg = "#d32f2f"
    button_hover = "#f3a5a5"

    # Styles
    st.markdown(f"""
    <style>
        .stApp {{
            padding-top: 0rem !important;
        }}
                
        .block-container {{
            padding-top: 2rem !important;
            padding-left: 2rem !important;
            padding-right: 2rem !important;
        }}
        h1, h2, h3, h4, h5, h6 {{
            color: {text_color} !important;
        }}
        html, body, [class*="css"], label {{
            color: {text_color} !important;
            font-family: 'Arial', sans-serif;
        }}
        [data-testid="stAppViewContainer"], [data-testid="stAppViewContainer"] > .main {{
            background-color: {bg_color} !important;
            color: {text_color} !important;
        }}

        /* Text input */
        .stTextInput > div {{
            background-color: {input_bg} !important;
            border-radius: 25px !important;
            padding: 2px !important;
            border: 1px solid {input_border} !important;
        }}
        .stTextInput input {{
            background-color: {input_bg} !important;
            border: none !important;
            color: {text_color} !important;
            padding-left: 10px !important;
            font-size: 1rem !important;
            caret-color: {"black" if is_light else "white"} !important;
        }}
        .stTextInput input::placeholder {{
            color: {placeholder_color} !important;
        }}

        div.stTextInput > div > div > div[data-testid="InputInstructions"],
        [data-testid="InputInstructions"] {{
            display: none !important;
        }}

        /* Date input */
        div[data-testid="stDateInput"] > div {{
            background-color: {input_bg} !important;
            border-radius: 25px !important;
            border: 1px solid {input_border} !important;
            padding: 0 10px !important;
        }}
        div[data-testid="stDateInput"] > div > div {{
            background-color: transparent !important;
            border: none !important;
            box-shadow: none !important;
        }}
        div[data-testid="stDateInput"] input {{
            background-color: {input_bg} !important;
            border: none !important;
            color: {text_color} !important;
            padding: 8px 0 !important;
            box-shadow: none !important;
        }}
        div[data-testid="stDateInput"] svg {{
            color: {text_color} !important;
        }}

        div[data-baseweb="popover"] {{
            background-color: {calendar_dropdown_bg} !important;
            border-radius: 15px !important;
            border: 1px solid {calendar_border} !important;
            box-shadow: none !important;
        }}
        
        div[data-testid="stDateInput"] > div:focus-within {{
            box-shadow: none !important;
        }}

        /* Selectboxes */
        .stSelectbox div[data-baseweb="select"] > div,
        .stSelectbox div[role="listbox"],
        .stSelectbox li {{
            background-color: {select_bg} !important;
            color: {text_color} !important;
            border: 1px solid {input_border} !important;
            border-radius: 25px !important;
        }}
        .stSelectbox li:hover {{
            background-color: {select_hover} !important;
        }}

        /* Download buttons */
        div.stDownloadButton > button {{
            background-color: {button_bg} !important;
            color: white !important;
            font-weight: bold !important;
            border: none !important;
            border-radius: 25px !important;
            padding: 8px 20px !important;
            font-size: 1rem !important;
            margin-left: -20px;
            margin-top: 5px;
            min-width: 120px !important;
            max-width: 180px !important;
            white-space: nowrap !important;
        }}

        div.stDownloadButton > button:hover {{
            background-color: {button_hover} !important;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.2) !important;
        }}

        /* Custom button styling ONLY for main content, NOT sidebar */
        .block-container div[data-testid="stButton"] > button,
        .block-container div[data-testid^="FormSubmitter"] button {{
            background-color: {button_bg} !important;
            color: white !important;
            font-weight: bold !important;
            border: none !important;
            border-radius: 25px !important;
            padding: 10px 20px !important;
            font-size: 1rem !important;
            margin-left: -5px;
            margin-top: 6px;
            min-width: 120px !important;
            max-width: 180px !important;
            white-space: nowrap !important;
        }}

        .block-container div[data-testid="stButton"] > button:hover,
        .block-container div[data-testid^="FormSubmitter"] button:hover {{
            background-color: {button_hover} !important;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.2) !important;
        }}

        /* Full cleanup of folium iframe borders & background */
        iframe, .folium-map {{
            background-color: transparent !important;
            margin: 0 !important;
            padding: 0 !important;
            border: none !important;
            box-shadow: none !important;
            outline: none !important;
        }}

        /* Clear surrounding Streamlit container spacing */
        .element-container,
        .st-emotion-cache-1v0mbdj,
        .st-emotion-cache-ocqkz7,
        .st-emotion-cache-1kyxreq {{
            background-color: transparent !important;
            margin: 0 !important;
            padding: 0 !important;
            box-shadow: none !important;
            border: none !important;
        }}

        div[data-testid="stVerticalBlock"] {{
            padding: 0 !important;
            margin: 0 !important;
            box-shadow: none !important;
        }}
        div[data-testid="stAlert"] * {{
            color: {error_text} !important;
            font-weight: 500 !important;
        }}

        html, body {{
            overflow-x: hidden !important;
            background: transparent !important;
        }}
        .folium-map {{
        animation: fadeInMap 0.1s ease-in;
    }}

    @keyframes fadeInMap {{
        from {{ opacity: 0; }}
        to {{ opacity: 1; }}
    }}

    </style>
    """, unsafe_allow_html=True)

    # Row 2: Filters - Date (Month/Year), Barangay, Age, Sex, and Reset Button
    col_date, col_barangay, col_age_group, col_sex, col_reset = st.columns([1.5, 0.8, 0.8, 0.8, 0.5])

    with col_date:

        col_month, col_year = st.columns([1, 1])

        if "selected_month" not in st.session_state.heatmap_filters:
            st.session_state.heatmap_filters["selected_month"] = datetime.today().month
        if "selected_year" not in st.session_state.heatmap_filters:
            st.session_state.heatmap_filters["selected_year"] = datetime.today().year
        
        with col_month:
            # Check if current stored value is "All" or a month number
            st.markdown('<div class="filter-label">Month</div>', unsafe_allow_html=True)
            current_stored = st.session_state.heatmap_filters.get("selected_month", "All") 
            month_options = ["All"] + list(range(1, 13))
            new_month = st.selectbox(
                "",
                options=month_options,
                format_func=lambda x: "All" if x == "All" else datetime(2000, x, 1).strftime('%B'),
                index=month_options.index(current_stored) if current_stored in month_options else 0,
                label_visibility="collapsed",
                key="month_input"
            )
        
        with col_year:
            st.markdown('<div class="filter-label">Year</div>', unsafe_allow_html=True)
            current_year = st.session_state.heatmap_filters.get("selected_year", datetime.today().year)
            year_range = list(range(2020, datetime.today().year + 2))
            new_year = st.selectbox(
                "",
                options=year_range,
                index=year_range.index(current_year) if current_year in year_range else len(year_range) - 1,
                label_visibility="collapsed",
                key="year_input"
            )
        
        # Update session state if month or year changed
        if (new_month != st.session_state.heatmap_filters["selected_month"] or
            new_year != st.session_state.heatmap_filters["selected_year"]):
            st.session_state.heatmap_filters["selected_month"] = new_month
            st.session_state.heatmap_filters["selected_year"] = new_year
            st.rerun()

    with col_barangay:
        st.markdown('<div class="filter-label">Barangay</div>', unsafe_allow_html=True)
        new_barangay = st.selectbox(
            "", [
                "All", "Alang-Alang", "Bakilid", "Banilad", "Basak", "Cabancalan", "Cambaro", "Canduman",
                "Casili", "Casuntingan", "Centro (Poblacion)", "Cubacub", "Guizo", "Ibabao-Estancia", "Jagobiao",
                "Labogon", "Looc", "Maguikay", "Mantuyong", "Opao", "Pagsabungan", "Pakna-an", "Subangdaku",
                "Tabok", "Tawason", "Tingub", "Tipolo", "Umapad"
            ],
            index=[
                "All", "Alang-Alang", "Bakilid", "Banilad", "Basak", "Cabancalan", "Cambaro", "Canduman",
                "Casili", "Casuntingan", "Centro (Poblacion)", "Cubacub", "Guizo", "Ibabao-Estancia", "Jagobiao",
                "Labogon", "Looc", "Maguikay", "Mantuyong", "Opao", "Pagsabungan", "Pakna-an", "Subangdaku",
                "Tabok", "Tawason", "Tingub", "Tipolo", "Umapad"
            ].index(st.session_state.heatmap_filters["barangay"]),
            label_visibility="collapsed", 
            key="barangay_input"
        )
        if new_barangay != st.session_state.heatmap_filters["barangay"]:
            st.session_state.heatmap_filters["barangay"] = new_barangay
            st.rerun()

    with col_age_group:
        st.markdown('<div class="filter-label">Age Group</div>', unsafe_allow_html=True)
        age_group_options = ["All", "0-14", "15-24", "25-64", "65+"]

        current_age_group = st.session_state.heatmap_filters.get("age_group", "All")
        new_age_group = st.selectbox(
            "",
            options=age_group_options,
            index=age_group_options.index(current_age_group) if current_age_group in age_group_options else 0,
            label_visibility="collapsed",
            key="age_group_input"
        )

        if new_age_group != st.session_state.heatmap_filters["age_group"]:
            st.session_state.heatmap_filters["age_group"] = new_age_group
            st.rerun()

    with col_sex:
        st.markdown('<div class="filter-label">Sex</div>', unsafe_allow_html=True)
        new_sex = st.selectbox(
            "", 
            ["All", "Male", "Female"], 
            index=["All", "Male", "Female"].index(st.session_state.heatmap_filters["sex"]),
            label_visibility="collapsed", 
            key="sex_input"
        )
        if new_sex != st.session_state.heatmap_filters["sex"]:
            st.session_state.heatmap_filters["sex"] = new_sex
            st.rerun()

    with col_reset:
        st.markdown('<div class="reset-btn-container">', unsafe_allow_html=True)
        if st.button("Reset Filters", key="reset_filters"):
            st.session_state.heatmap_filters = {
                "date": datetime.today().date(),
                "barangay": "All",
                "age_group": "All",
                "sex": "All"
            }
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)


    # Use the filters from session state in the query
    heatmap_date_filter = st.session_state.heatmap_filters["date"]
    heatmap_barangay_filter = st.session_state.heatmap_filters["barangay"]
    heatmap_sex_filter = st.session_state.heatmap_filters["sex"]

    # --- Fetch and Filter Records from Supabase ---
    query = supabase.table("HEATMAP_Table")\
    .select("*, DIAGNOSIS_Table(DX_ID, DX_STATUS)")

    # Apply filters only if they're not set to "All" or None
    if heatmap_barangay_filter != "All":
        query = query.ilike("MAP_BRGY", f"%{heatmap_barangay_filter}%")

    if heatmap_sex_filter != "All":
        query = query.eq("MAP_SEX", heatmap_sex_filter)

    # Handle month/year filter
    selected_month_value = st.session_state.heatmap_filters.get("selected_month", "All")
    selected_year = st.session_state.heatmap_filters.get("selected_year", datetime.today().year)

    if selected_month_value != "All":
        # First and last day of chosen month
        start_date = datetime(selected_year, selected_month_value, 1)
        if selected_month_value == 12:
            end_date = datetime(selected_year + 1, 1, 1) - timedelta(seconds=1)
        else:
            end_date = datetime(selected_year, selected_month_value + 1, 1) - timedelta(seconds=1)
        
        query = query.gte("MAP_GENERATED_AT", start_date.isoformat())
        query = query.lte("MAP_GENERATED_AT", end_date.isoformat())
    else:
        # If month is "All", filter by year only
        start_date = datetime(selected_year, 1, 1)
        end_date = datetime(selected_year + 1, 1, 1) - timedelta(seconds=1)
        query = query.gte("MAP_GENERATED_AT", start_date.isoformat())
        query = query.lte("MAP_GENERATED_AT", end_date.isoformat())

    # Get age group filter
    age_group_filter = st.session_state.heatmap_filters.get("age_group", "All")

    # Execute query
    try:
        res = query.execute()
        heatmap_records = res.get("data", []) if isinstance(res, dict) else getattr(res, "data", [])
        
        confirmed_counts = defaultdict(int)

        for record in heatmap_records:
            brgy = record.get("MAP_BRGY", "")
            date = record.get("MAP_GENERATED_AT", "")
            
            try:
                dt = parse(date)
                # Check if the record should be processed based on month filter
                if selected_month_value == "All":
                    if dt.year == selected_year:
                        process_record = True
                    else:
                        continue
                else:
                    if dt.month == selected_month_value and dt.year == selected_year:
                        process_record = True
                    else:
                        continue
                
                if process_record:
                    diagnosis_info = record.get("DIAGNOSIS_Table", {})
                    if isinstance(diagnosis_info, list):
                        diagnosis_info = diagnosis_info[0] if diagnosis_info else {}
                    
                    dx_status = diagnosis_info.get("DX_STATUS")

                    if dx_status == "Confirmed Positive":
                        if age_group_filter != "All":
                            try:
                                age_group = record.get("MAP_AGE_GROUP", "")
                                # Convert age group string to numeric range for easier comparison
                                if age_group == "0-14":
                                    age_match = age_group_filter == "0-14"
                                elif age_group == "15-24":
                                    age_match = age_group_filter == "15-24"
                                elif age_group == "25-64":
                                    age_match = age_group_filter == "25-64"
                                elif age_group == "65+":
                                    age_match = age_group_filter == "65+"
                                else:
                                    age_match = False
                            except:
                                age_match = False
                            
                            if not age_match:
                                continue  # Skip this record if age group doesn't match
                        
                        # Only count if it passed all filters
                        confirmed_counts[brgy] += 1

            except Exception as e:
                print(f"Error processing record: {e}")
                continue

        # --- Compute Target & Identify High Risk ---

        risk_levels = {}
        targets = {}

        for brgy, pop in mandaue_barangay_population_2025.items():
            count = confirmed_counts.get(brgy, 0)

            # Determine if user selected full year or a specific month
            month_index = 0 if selected_month_value == "All" else 1

            # --- Target calculation ---
            if month_index == 0:
                target = pop * incidence_rate  # Annual target
                target_label = "Annual Target"
            else:
                target = (pop * incidence_rate) / 12  # Monthly target
                target_label = "Monthly Target"

            # Round target for display & risk calculation
            target_display = int(round(target))
            targets[brgy] = target_display  # store for popup
            target = target_display

            # How many cases were observed relative to the target
            risk_ratio = count / target if target > 0 else 0

            # Converts the ratio into a percentage, avoids decimals, caps at 100%
            risk_percent = min(round(risk_ratio * 100), 100)

            # Risk level classification based on risk_percent
            if risk_percent >= 75:
                color = "red"
                risk_level = "High"
            elif risk_percent >= 55:
                color = "orange"
                risk_level = "Moderate"
            else:
                color = "green"
                risk_level = "Low"

            risk_levels[brgy] = (risk_level + " Risk", color)

    except Exception as e:
        st.error(f"Failed to fetch heatmap data: {e}")
        heatmap_records = []
        risk_levels = {}

   # --- Prepare Heatmap ---
    heat_data = []
    confirmed_records = []  # Store only confirmed records for mapping

    for record in heatmap_records:
        # Check if this record represents a confirmed case
        diagnosis_info = record.get("DIAGNOSIS_Table", {})
        if isinstance(diagnosis_info, list):
            diagnosis_info = diagnosis_info[0] if diagnosis_info else {}
        
        dx_status = diagnosis_info.get("DX_STATUS")
        
        # Only process confirmed positive cases
        if dx_status == "Confirmed Positive":
            if age_group_filter != "All":
                try:
                    age_group = record.get("MAP_AGE_GROUP", "")
                    if age_group == "0-14":
                        age_match = age_group_filter == "0-14"
                    elif age_group == "15-24":
                        age_match = age_group_filter == "15-24"
                    elif age_group == "25-64":
                        age_match = age_group_filter == "25-64"
                    elif age_group == "65+":
                        age_match = age_group_filter == "65+"
                    else:
                        age_match = False
                except:
                    age_match = False
                
                if not age_match:
                    continue
            
            lat = record.get("MAP_LAT")
            lon = record.get("MAP_LANG")
            brgy = record.get("MAP_BRGY", "")
            
            if lat is None or lon is None:
                continue
            
            confirmed_records.append(record)  # Add to confirmed records list
            
            # Get the risk level and assign a weight
            risk_label, risk_color = risk_levels.get(brgy, ("Unknown", "gray"))
            if risk_color == "red":
                weight = 3  # Highest intensity for high risk
            elif risk_color == "orange":
                weight = 2  # Medium intensity for moderate risk
            else:
                weight = 1  # Lowest intensity for low risk
            
            heat_data.append([lat, lon, weight])

    m = folium.Map(location=[10.3200, 123.9000], zoom_start=13, tiles='OpenStreetMap')

    if heat_data:
        custom_gradient = {
            0.3: 'green',   # Low risk
            0.6: 'orange',  # Moderate risk
            1.0: 'red'     # High risk
        }
        
        HeatMap(
            heat_data,
            radius=20, 
            blur=15,   
            min_opacity=0.5,
            max_zoom=18,
            gradient=custom_gradient
        ).add_to(m)

    # Add barangay-level summary markers (larger and more opaque)
    for brgy, (risk_label, risk_color) in risk_levels.items():
        # Only show barangays that have confirmed cases
        if confirmed_counts.get(brgy, 0) == 0:
            continue
            
        # Find first confirmed record for this barangay to get coordinates
        brgy_confirmed_records = [r for r in confirmed_records if r.get("MAP_BRGY", "") == brgy]
        if not brgy_confirmed_records:
            continue
            
        lat = brgy_confirmed_records[0].get("MAP_LAT")
        lon = brgy_confirmed_records[0].get("MAP_LANG")
        if lat is None or lon is None:
            continue
            
        target_value = int(round(targets.get(brgy, 0)))
        confirmed_cases = confirmed_counts.get(brgy, 0)
        population = mandaue_barangay_population_2025.get(brgy, 0)
        
        popup_content = f"""
        <div style="font-family: Arial; min-width: 200px">
            <h4>{brgy} Summary</h4>
            <hr style="margin: 5px 0;">
            <b>Population:</b> {population}<br>
            <b>Confirmed Cases:</b> {confirmed_cases}<br>
            <b>{target_label}:</b> {target_value}<br>
            <b>Risk Level:</b> <span style="color:{risk_color}; font-weight:bold;">{risk_label}</span>
        </div>
        """
        
        folium.CircleMarker(
            location=[lat, lon],
            radius=10,
            color=risk_color,
            weight=2,
            fill=True,
            fill_color=risk_color,
            fill_opacity=0.7,
            popup=folium.Popup(popup_content, max_width=300)
        ).add_to(m)

    # --- Floating Match Count Badge ---
    confirmed_count = len(confirmed_records)
    badge_html = f"""
    <div style="
        position: absolute;
        top: 10px;
        left: 50%;
        transform: translateX(-50%);
        background-color: {'#ee5a5a' if confirmed_count > 0 else '#58a83e'};
        color: white;
        border-radius: 35px;
        height: 45px;
        padding: 12px 20px 10px 22px;
        font-weight: bold;
        font-size: 15px;
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
        z-index: 9999;
        text-align: center;
    ">
        🔍 {confirmed_count} case{'s' if confirmed_count != 1 or 0 else ''} found
    </div>
    """

    # Inject into the Folium map so it's included in HTML export
    m.get_root().html.add_child(Element(badge_html))

    # --- Display Map ---
    map_html = m.get_root().render()
    safe_map_html = map_html.replace('"', '&quot;')

    components.html(
        f"""
        <div style="display: flex; justify-content: center; align-items: center; width: 100%; margin: 0 auto; padding: 0;">
            <iframe srcdoc="{safe_map_html}" width="1200" height="650" style="
                border: none;
                margin: 0;
                padding: 0;
                background: transparent;
                overflow: hidden;
                border-radius: 15px;
            "></iframe>
        </div>
        """,
        height=660,
    )

    # --- Downloads ---
    if confirmed_records: # Only show if there are records
        selected_month_value = st.session_state.heatmap_filters.get("selected_month", "All")
        selected_year_value = st.session_state.heatmap_filters.get("selected_year", datetime.today().year)
        excel_bytes = generate_heatmap_excel(confirmed_records, risk_levels, selected_month_value, selected_year_value)
        
        col_left, btn_col1, spacer_col, btn_col2 = st.columns([8, 1, 0.3, 1])
        with col_left:
            st.write("")
        with btn_col1: # HTML download button
            st.download_button(
                "Export Map", 
                map_html, 
                f"TB_Cases_Heatmap_{datetime.now().strftime('%Y-%m-%d')}.html", 
                "text/html", 
                key="download-image"
            )
        with spacer_col:
            st.write("")
        with btn_col2: # Excel download button
            st.download_button(
                "Export Excel", 
                data=excel_bytes, 
                file_name=f"TB_Cases_Heatmap_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                key="download-excel"
            )

    st.markdown('<div style="height:10px;"></div>', unsafe_allow_html=True)