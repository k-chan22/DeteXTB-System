# Reports.py

import streamlit as st
import io
from Supabase import supabase
from datetime import datetime
from collections import Counter
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, Reference
import matplotlib.pyplot as plt
from io import BytesIO


# --- Data Fetching Helper Function Logic + Constant ---

# Get today's date in YYYY-MM-DD format
today_str = datetime.now().strftime("%Y-%m-%d")

# Function to fetch the logic for the AI Presumptive TB Report block
def fetch_ai_report_data(selected_month=None, selected_year=None):
    # Fetch AI-predicted results with associated chest X-ray IDs (CXR_ID)
    ai_results_resp = supabase.table("RESULT_Table").select("CXR_ID", "RES_PRESUMPTIVE", "RES_DATE", "RES_CONF_SCORE").execute()
    ai_results = ai_results_resp.data or []

    if not ai_results:
        return {
            "Total Flagged Patients": "0",
            "AI Accuracy Rate": "No Presumptive Cases",
            "Pending Confirmations": "0"
        }

    # Filter by month/year if provided
    if selected_month != "All" or selected_year:
        filtered_results = []
        for item in ai_results:
            if item.get('RES_DATE'):
                try:
                    result_date = datetime.fromisoformat(item['RES_DATE'].replace('Z', '+00:00'))
                    if selected_year and result_date.year != selected_year:
                        continue
                    if selected_month != "All" and result_date.month != selected_month:
                        continue
                    filtered_results.append(item)
                except (ValueError, TypeError):
                    continue
        ai_results = filtered_results

    if not ai_results:
        return {
            "Total Flagged Patients": "0",
            "AI Accuracy Rate": "No Presumptive Cases",
            "Pending Confirmations": "0"
        }

    flagged_cxr_ids = [item['CXR_ID'] for item in ai_results if str(item['RES_PRESUMPTIVE']).lower() == "positive"]

    # Get unique patient IDs
    total_flagged = "0"
    if flagged_cxr_ids:
        cxr_resp = supabase.table("CHEST_XRAY_Table").select("CXR_ID", "PT_ID").in_("CXR_ID", flagged_cxr_ids).execute()
        pt_ids = [item['PT_ID'] for item in cxr_resp.data or []]
        total_flagged = str(len(set(pt_ids)))

    # Fetch diagnosis statuses
    cxr_ids_all = [item['CXR_ID'] for item in ai_results]
    diagnosis_dict = {}
    if cxr_ids_all:
        dx_resp = supabase.table("DIAGNOSIS_Table").select("CXR_ID", "DX_STATUS", "DX_UPDATED_AT").in_("CXR_ID", cxr_ids_all).execute()
        diagnosis_dict = {item['CXR_ID']: {"status": item['DX_STATUS'], "updated_at": item.get('DX_UPDATED_AT')} for item in dx_resp.data or []}

    # Pending confirmations
    pending_count = str(sum(
        1 for cxr_id in flagged_cxr_ids
        if diagnosis_dict.get(cxr_id) is None or str(diagnosis_dict[cxr_id]["status"]).strip().lower() == "pending"
    ))

    # Normalize and calculate AI accuracy
    def normalize_dx_status(dx_status):
        if not dx_status:
            return None
        dx_status = str(dx_status).lower()
        if "confirmed positive" in dx_status:
            return "positive"
        elif "confirmed negative" in dx_status:
            return "negative"
        elif dx_status == "pending":
            return "pending"
        return None

    correct = 0
    evaluated = 0
    for item in ai_results:
        cxr_id = item['CXR_ID']
        ai_pred = str(item['RES_PRESUMPTIVE']).lower()
        dx_info = diagnosis_dict.get(cxr_id)
        dx_status = normalize_dx_status(dx_info["status"]) if dx_info else None

        if dx_status and dx_status != "pending":
            evaluated += 1
            if ai_pred == dx_status:
                correct += 1

    accuracy = f"{(correct / evaluated) * 100:.2f}%" if evaluated else "No Presumptive Cases"

    return {
        "Total Flagged Patients": total_flagged,
        "AI Accuracy Rate": str(accuracy),
        "Pending Confirmations": pending_count
    }

# Function to fetch detailed flagged patient data for AI report exports
def fetch_flagged_patient_details(selected_month=None, selected_year=None):
    # Fetch AI results with confidence scores (level is used)
    ai_results_resp = supabase.table("RESULT_Table").select("CXR_ID", "RES_PRESUMPTIVE", "RES_DATE", "RES_CONF_SCORE").execute()
    ai_results = ai_results_resp.data or []
    
    # Filter by date
    if selected_month != "All" or selected_year:
        filtered_results = []
        for item in ai_results:
            if item.get('RES_DATE'):
                try:
                    result_date = datetime.fromisoformat(item['RES_DATE'].replace('Z', '+00:00'))
                    if selected_year and result_date.year != selected_year:
                        continue
                    if selected_month != "All" and result_date.month != selected_month:
                        continue
                    filtered_results.append(item)
                except (ValueError, TypeError):
                    continue
        ai_results = filtered_results
    
    # Get only flagged (positive) cases
    flagged_results = [item for item in ai_results if str(item['RES_PRESUMPTIVE']).lower() == "positive"]
    
    if not flagged_results:
        return []
    
    # Get CXR IDs
    flagged_cxr_ids = [item['CXR_ID'] for item in flagged_results]
    
    # Fetch patient info with additional fields
    cxr_resp = supabase.table("CHEST_XRAY_Table").select("CXR_ID", "PT_ID").in_("CXR_ID", flagged_cxr_ids).execute()
    cxr_to_patient = {item['CXR_ID']: item['PT_ID'] for item in cxr_resp.data or []}
    
    # Fetch patient details with Age, Sex, and Barangay
    patient_ids = list(set(cxr_to_patient.values()))
    patient_details_dict = {}
    if patient_ids:
        patient_resp = supabase.table("PATIENT_Table").select("PT_ID", "PT_AGE", "PT_SEX", "PT_BRGY").in_("PT_ID", patient_ids).execute()
        for patient in patient_resp.data or []:
            # Normalize sex to 'M' or 'F' to save space
            raw_sex = str(patient.get('PT_SEX', '')).strip().lower()
            if raw_sex.startswith('m'):
                sex_value = 'M'
            elif raw_sex.startswith('f'):
                sex_value = 'F'
            else:
                sex_value = 'N/A'

            patient_details_dict[patient['PT_ID']] = {
                "age": patient.get('PT_AGE', 'N/A'),
                "sex": sex_value,
                "barangay": patient.get('PT_BRGY', 'N/A')
            }
    
    # Fetch diagnosis info
    dx_resp = supabase.table("DIAGNOSIS_Table").select("CXR_ID", "DX_STATUS", "DX_UPDATED_AT").in_("CXR_ID", flagged_cxr_ids).execute()
    diagnosis_dict = {item['CXR_ID']: {"status": item['DX_STATUS'], "updated_at": item.get('DX_UPDATED_AT')} for item in dx_resp.data or []}
    
    # Build detailed patient list with additional fields
    patient_details = []
    for result in flagged_results:
        cxr_id = result['CXR_ID']
        pt_id = cxr_to_patient.get(cxr_id, "Unknown")
        dx_info = diagnosis_dict.get(cxr_id)
        patient_info = patient_details_dict.get(pt_id, {})
        
        # Format final status
        final_status = "Pending"
        confirmation_date = "N/A"
        if dx_info:
            status_raw = str(dx_info["status"]).lower()
            if "confirmed positive" in status_raw:
                final_status = "Confirmed Positive"
            elif "confirmed negative" in status_raw:
                final_status = "Confirmed Negative"
            elif "pending" in status_raw:
                final_status = "Pending"
            
            if dx_info["updated_at"]:
                try:
                    confirmation_date = datetime.fromisoformat(dx_info["updated_at"].replace('Z', '+00:00')).strftime("%Y-%m-%d")
                except:
                    confirmation_date = "N/A"
        
        patient_details.append({
            "Patient ID": pt_id,
            "Age": patient_info.get('age', 'N/A'),
            "Sex": patient_info.get('sex', 'N/A'),
            "Barangay": patient_info.get('barangay', 'N/A'),
            "AI Flagged Date": result.get('RES_DATE', 'N/A'),
            "AI Confidence Score": f"{float(result.get('RES_CONF_SCORE', 0)) * 100:.1f}%" if result.get('RES_CONF_SCORE') else "N/A",
            "Final Status": final_status,
            "Confirmation Date": confirmation_date
        })
    
    return patient_details

# Function to calculate AI performance metrics (TP, TN, FP, FN)
def calculate_ai_performance_metrics(selected_month=None, selected_year=None):
    # Fetch all AI results
    ai_results_resp = supabase.table("RESULT_Table").select("CXR_ID", "RES_PRESUMPTIVE", "RES_DATE").execute()
    ai_results = ai_results_resp.data or []
    
    # Filter by date
    if selected_month != "All" or selected_year:
        filtered_results = []
        for item in ai_results:
            if item.get('RES_DATE'):
                try:
                    result_date = datetime.fromisoformat(item['RES_DATE'].replace('Z', '+00:00'))
                    if selected_year and result_date.year != selected_year:
                        continue
                    if selected_month != "All" and result_date.month != selected_month:
                        continue
                    filtered_results.append(item)
                except (ValueError, TypeError):
                    continue
        ai_results = filtered_results
    
    # Fetch diagnosis statuses
    cxr_ids_all = [item['CXR_ID'] for item in ai_results]
    diagnosis_dict = {}
    if cxr_ids_all:
        dx_resp = supabase.table("DIAGNOSIS_Table").select("CXR_ID", "DX_STATUS").in_("CXR_ID", cxr_ids_all).execute()
        diagnosis_dict = {item['CXR_ID']: item['DX_STATUS'] for item in dx_resp.data or []}
    
    # Normalize diagnosis status
    def normalize_dx_status(dx_status):
        if not dx_status:
            return None
        dx_status = str(dx_status).lower()
        if "confirmed positive" in dx_status:
            return "positive"
        elif "confirmed negative" in dx_status:
            return "negative"
        elif dx_status == "pending":
            return "pending"
        return None
    
    # Calculate metrics
    tp = tn = fp = fn = 0
    for item in ai_results:
        cxr_id = item['CXR_ID']
        ai_pred = str(item['RES_PRESUMPTIVE']).lower()
        dx_status = normalize_dx_status(diagnosis_dict.get(cxr_id))
        
        if dx_status and dx_status != "pending":
            if ai_pred == "positive" and dx_status == "positive":
                tp += 1
            elif ai_pred == "negative" and dx_status == "negative":
                tn += 1
            elif ai_pred == "positive" and dx_status == "negative":
                fp += 1
            elif ai_pred == "negative" and dx_status == "positive":
                fn += 1
    
    return {
        "True Positives (TP)": tp,
        "True Negatives (TN)": tn,
        "False Positives (FP)": fp,
        "False Negatives (FN)": fn,
        "Total Evaluated": tp + tn + fp + fn
    }

# Function to fetch the logic for the Confirmed TB Cases Report block
def fetch_confirmed_cases_data(selected_month=None, selected_year=None):
    # Fetch all confirmed status entries from RESULT_Table with dates
    result_resp = supabase.table("RESULT_Table").select("RES_STATUS", "RES_DATE").execute()
    result_data = result_resp.data or []

    # Filter by month/year
    if selected_month != "All" or selected_year:
        filtered_results = []
        for item in result_data:
            if item.get('RES_DATE'):
                try:
                    result_date = datetime.fromisoformat(item['RES_DATE'].replace('Z', '+00:00'))
                    if selected_year and result_date.year != selected_year:
                        continue
                    if selected_month != "All" and result_date.month != selected_month:
                        continue
                    filtered_results.append(item)
                except (ValueError, TypeError):
                    continue
        result_data = filtered_results

    # Normalize statuses
    statuses = [str(item['RES_STATUS']).strip().title() for item in result_data if item['RES_STATUS']]
    counter = Counter(statuses)

    positive = counter.get("Confirmed Positive", 0)
    negative = counter.get("Confirmed Negative", 0)
    total = positive + negative

    pos_percent = f"{(positive / total) * 100:.0f}%" if total else "0%"
    neg_percent = f"{(negative / total) * 100:.0f}%" if total else "0%"

    return {
        "Total Confirmed Cases": f"{total} Confirmed TB Cases",
        "Positive Cases": f"{positive} {pos_percent} of Total",
        "Negative Cases": f"{negative} {neg_percent} of Total"
    }

# Function to fetch detailed confirmed case data for exports
def fetch_confirmed_case_details(selected_month=None, selected_year=None):
    # Fetch confirmed results with dates
    result_resp = supabase.table("RESULT_Table").select("CXR_ID", "RES_STATUS", "RES_DATE").execute()
    result_data = result_resp.data or []
    
    # Filter by date and confirmed status
    confirmed_results = []
    for item in result_data:
        if not item.get('RES_STATUS') or 'confirmed' not in str(item['RES_STATUS']).lower():
            continue
        
        if item.get('RES_DATE'):
            try:
                result_date = datetime.fromisoformat(item['RES_DATE'].replace('Z', '+00:00'))
                if selected_year and result_date.year != selected_year:
                    continue
                if selected_month != "All" and result_date.month != selected_month:
                    continue
                confirmed_results.append(item)
            except (ValueError, TypeError):
                continue
    
    if not confirmed_results:
        return []
    
    # Get CXR IDs
    cxr_ids = [item['CXR_ID'] for item in confirmed_results]
    
    # Fetch patient info
    cxr_resp = supabase.table("CHEST_XRAY_Table").select("CXR_ID", "PT_ID").in_("CXR_ID", cxr_ids).execute()
    cxr_to_patient = {item['CXR_ID']: item['PT_ID'] for item in cxr_resp.data or []}
    
    # Fetch patient details with Age, Sex, and Barangay
    patient_ids = list(set(cxr_to_patient.values()))
    patient_details_dict = {}
    if patient_ids:
        patient_resp = supabase.table("PATIENT_Table").select("PT_ID", "PT_AGE", "PT_SEX", "PT_BRGY").in_("PT_ID", patient_ids).execute()
        for patient in patient_resp.data or []:
            # Normalize sex to 'M' or 'F' to save space
            raw_sex = str(patient.get('PT_SEX', '')).strip().lower()
            if raw_sex.startswith('m'):
                sex_value = 'M'
            elif raw_sex.startswith('f'):
                sex_value = 'F'
            else:
                sex_value = 'N/A'

            patient_details_dict[patient['PT_ID']] = {
                "age": patient.get('PT_AGE', 'N/A'),
                "sex": sex_value,
                "barangay": patient.get('PT_BRGY', 'N/A')
            }
    
    # Fetch diagnosis info (for confirmation method notes)
    dx_resp = supabase.table("DIAGNOSIS_Table").select("CXR_ID", "DX_NOTES").in_("CXR_ID", cxr_ids).execute()
    diagnosis_notes = {item['CXR_ID']: item.get('DX_NOTES', 'N/A') for item in dx_resp.data or []}
    
    # Fetch AI flagging info
    ai_resp = supabase.table("RESULT_Table").select("CXR_ID", "RES_PRESUMPTIVE").in_("CXR_ID", cxr_ids).execute()
    ai_flagged = {item['CXR_ID']: str(item.get('RES_PRESUMPTIVE', '')).lower() == "positive" for item in ai_resp.data or []}
    
    # Build detailed case list with additional fields
    case_details = []
    for result in confirmed_results:
        cxr_id = result['CXR_ID']
        pt_id = cxr_to_patient.get(cxr_id, "Unknown")
        patient_info = patient_details_dict.get(pt_id, {})
        
        # Determine final status
        status_raw = str(result['RES_STATUS']).lower()
        if "confirmed positive" in status_raw:
            final_status = "Confirmed Positive"
        elif "confirmed negative" in status_raw:
            final_status = "Confirmed Negative"
        else:
            final_status = result['RES_STATUS']
        
        # Extract confirmation method from notes
        notes = diagnosis_notes.get(cxr_id, 'N/A')
        confirmation_method = "Clinical Assessment"  # Default
        if notes and notes != 'N/A':
            # Look for keywords in notes
            notes_lower = notes.lower()
            if 'genexpert' in notes_lower or 'xpert' in notes_lower:
                confirmation_method = "GeneXpert"
            elif 'smear' in notes_lower:
                confirmation_method = "Smear Test"
            elif 'culture' in notes_lower:
                confirmation_method = "Culture"
            elif 'x-ray' in notes_lower or 'xray' in notes_lower:
                confirmation_method = "X-Ray Finding"
        
        case_details.append({
            "Patient ID": pt_id,
            "Age": patient_info.get('age', 'N/A'),
            "Sex": patient_info.get('sex', 'N/A'),
            "Barangay": patient_info.get('barangay', 'N/A'),
            "Final Status": final_status,
            "Confirmation Date": result.get('RES_DATE', 'N/A'),
            "Confirmation Method": confirmation_method,
            "AI Flagged": "Yes" if ai_flagged.get(cxr_id, False) else "No"
        })
    
    return case_details


# --- Export Helper Function Logic + Class ---

# Class to format PDF style
class PDFReport_format(FPDF):
    def header(self):
        try:
            # Left logo
            self.image("images/logoonly-dark.png", 10, 4, 40)  # (file, x, y, width)
        except RuntimeError as e:
            print("Left icon error:", e)
        try:
            # Right logo
            self.image("images/CTUlogo.png", 170, 10, 27)
        except RuntimeError as e:
            print("Right icon error:", e)

        self.ln(3)

        # First line: Mandaue City Health Office (MCHO)
        self.set_font('Arial', 'B', 12)
        self.cell(0, 6, "Mandaue City Health Office", align='C', ln=1)

        # Next lines: MCHO Address and Contact Info
        self.set_font('Arial', '', 10)
        self.cell(0, 6, "S.B. Cabahug, Mandaue City, Philippines.", align='C', ln=1)
        self.cell(0, 6, "Call us on: +63 (032) 230 4500 | FB: Mandaue City Public Affairs Office |", align='C', ln=1)
        self.cell(0, 6, "Email: cmo@mandauecity.gov.ph", align='C', ln=1)

        self.ln(10) # Spacer

        # System name
        self.set_font('Arial', 'B', 12)
        self.cell(0, 6, "DeteXTB: AI-Assisted Presumptive Tuberculosis Detection and Mapping System", align='C', ln=1)
        self.ln(6)

    # Function to format the metrics and values
    def write_key_value(self, key, value, value_color=(0,0,0), italic_in_parens=False):
        self.set_text_color(*value_color)
        full_text = f"{key}: {value}"

        if italic_in_parens and "(" in full_text and ")" in full_text:
            start = full_text.find("(")
            end = full_text.find(")") + 1

            before_paren = full_text[:start]
            paren_text = full_text[start:end]
            after_paren = full_text[end:]

            # Use consistent font size for all parts
            font_size = 13

            # Before parentheses format
            self.set_font("Arial", "B", font_size)
            self.cell(self.get_string_width(before_paren), 10, before_paren, ln=0)

            # Parentheses format
            self.set_font("Arial", "I", font_size)
            self.cell(self.get_string_width(paren_text), 10, paren_text, ln=0)

            # After parentheses format
            if after_paren.strip():
                self.set_font("Arial", "", font_size)
                self.cell(self.get_string_width(after_paren), 10, after_paren, ln=0)

            self.ln(10)
        else:
            # If no parentheses or italic flag not set, print full text in bold
            self.set_font("Arial", "B", 13)
            self.cell(0, 10, full_text, ln=1)

# Function to generate AI Presumptive TB Report PDF
def generate_ai_pdf(report_title, data, filter_info=None, selected_month=None, selected_year=None):
    pdf = PDFReport_format()
    pdf.add_page()

    # Report title
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, report_title, ln=1, align="C")
    pdf.ln(5)

    # Add filter information if provided
    if filter_info:
        pdf.set_font("Arial", "I", 10)
        pdf.cell(0, 8, filter_info, ln=1, align="C")
        pdf.ln(5)

    # Draw horizontal line
    pdf.set_draw_color(0, 0, 0)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    # Color Format
    dark_orange = (204, 122, 42)
    red = (217, 57, 53)
    green = (0, 128, 0)
    black = (0, 0, 0)

    # Summary section
    for key, value in data.items():
        italicize = "(" in value and ")" in value

        if "True Negative" in key:
            value 

        if "Pending" in key:
            color = dark_orange
        elif "Positive" in key:
            color = red
        elif "Negative" in key:
            color = green
        else:
            color = black

        pdf.write_key_value(key, value, value_color=color, italic_in_parens=italicize)

    pdf.ln(5)

    # AI Accuracy Explanation Section
    pdf.set_font("Arial", "B", 12)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 10, "AI Performance Analysis", ln=1)
    pdf.set_font("Arial", "", 10)

    metrics = calculate_ai_performance_metrics(selected_month, selected_year)
    if metrics["Total Evaluated"] > 0:
        explanation = (
            f"The AI accuracy rate is calculated based on {metrics['Total Evaluated']} clinically reviewed outcomes "
            f"during this period: {metrics['True Positives (TP)']} True Positive(s), "
            f"{metrics['True Negatives (TN)']} True Negative(s), "
            f"{metrics['False Positives (FP)']} False Positive(s), "
            f"and {metrics['False Negatives (FN)']} False Negative(s). "
            "Negative counts are included for reference only."
        )
    else:
        explanation = (
            "No clinically reviewed outcomes are available for accuracy calculation during this period."
        )

    pdf.multi_cell(0, 6, explanation, align="J")
    pdf.ln(5)

    # Flagged Patient Details Section
    patient_details = fetch_flagged_patient_details(selected_month, selected_year)
    if patient_details:
        try:
            patient_details.sort(
                key=lambda x: (
                    datetime.strptime(str(x.get("AI Flagged Date", ""))[:10], "%Y-%m-%d")
                    if x.get("AI Flagged Date") else datetime.max,
                    str(x.get("Patient ID", "")).lower()
                )
            )
        except Exception as e:
            print("Sorting error (AI Presumptive):", e)

    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Flagged Patient Details", ln=1)

    if patient_details:
        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(200, 200, 200)
        
        col_widths = [8, 28, 10, 12, 25, 25, 24, 25, 33]
        headers = ["#", "Patient ID", "Age", "Sex", "Barangay", "Flagged Date", "AI Confidence", "Final Status", "Confirmation Date"]

        total_table_width = sum(col_widths)
        if total_table_width > 190:
            scale_factor = 190 / total_table_width
            col_widths = [int(w * scale_factor) for w in col_widths]

        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 8, header, border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Arial", "", 7) 
        for idx, patient in enumerate(patient_details, start=1):
            status = str(patient["Final Status"]).strip().lower()

            if "positive" in status:
                pdf.set_fill_color(255, 150, 150)
                fill = True
            else:
                fill = False

            confirm_date = str(patient.get("Confirmation Date", "")).strip()
            if not confirm_date or confirm_date.upper() == "N/A":
                confirm_date = "Not yet confirmed"
            else:
                confirm_date = confirm_date[:10]

            patient_id = str(patient["Patient ID"])
            barangay = str(patient["Barangay"])
            final_status = str(patient["Final Status"])

            pdf.cell(col_widths[0], 8, str(idx), border=1, align="C", fill=fill)
            pdf.cell(col_widths[1], 8, patient_id, border=1, align="C", fill=fill)
            pdf.cell(col_widths[2], 8, str(patient["Age"]), border=1, align="C", fill=fill)
            pdf.cell(col_widths[3], 8, str(patient["Sex"]), border=1, align="C", fill=fill)
            pdf.cell(col_widths[4], 8, barangay, border=1, align="C", fill=fill)
            pdf.cell(col_widths[5], 8, str(patient["AI Flagged Date"])[:10], border=1, align="C", fill=fill)
            pdf.cell(col_widths[6], 8, str(patient["AI Confidence Score"]), border=1, align="C", fill=fill)
            pdf.cell(col_widths[7], 8, final_status, border=1, align="C", fill=fill)
            pdf.cell(col_widths[8], 8, confirm_date, border=1, align="C", fill=fill)
            pdf.ln()

        # Demographic Summary Section
        pdf.ln(8)
        
        # Calculate demographic summaries
        age_groups = {
            "Children (0-14)": 0,
            "Youth/Young Adults (15-24)": 0,
            "Adults (25-64)": 0,
            "Elderly (65+)": 0
        }
        
        sex_counts = {}
        barangay_counts = {}
        
        for patient in patient_details:
            # Age group calculation
            age = patient.get("Age")
            if age and str(age).isdigit():
                age_int = int(age)
                if age_int <= 14:
                    age_groups["Children (0-14)"] += 1
                elif age_int <= 24:
                    age_groups["Youth/Young Adults (15-24)"] += 1
                elif age_int <= 64:
                    age_groups["Adults (25-64)"] += 1
                else:
                    age_groups["Elderly (65+)"] += 1
            
            # Sex count
            raw_sex = patient.get("Sex", "Unknown").strip().upper()
            if raw_sex == "F":
                sex = "Female"
            elif raw_sex == "M":
                sex = "Male"
            else:
                sex = "Unknown"

            sex_counts[sex] = sex_counts.get(sex, 0) + 1
            
            # Barangay count
            barangay = patient.get("Barangay", "Unknown")
            barangay_counts[barangay] = barangay_counts.get(barangay, 0) + 1
        
        # all_barangays = sorted(barangay_counts.items(), key=lambda x: x[0])  # alphabetical
        all_barangays = sorted(barangay_counts.items(), key=lambda x: x[1], reverse=True) # descending
        
        # Age Group Distribution Table
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, "Age Group Distribution", ln=1)
        
        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(60, 8, "Age Group", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Count", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Percentage", border=1, align="C", fill=True)
        pdf.ln()
        
        pdf.set_font("Arial", "", 8)
        total_patients = len(patient_details)
        for age_group, count in age_groups.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            pdf.cell(60, 8, age_group, border=1, align="L")
            pdf.cell(30, 8, str(count), border=1, align="C")
            pdf.cell(30, 8, f"{percentage:.1f}%", border=1, align="C")
            pdf.ln()
        
        pdf.ln(5)
        
        # Sex Distribution Table
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, "Sex Distribution", ln=1)
        
        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(40, 8, "Sex", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Count", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Percentage", border=1, align="C", fill=True)
        pdf.ln()
        
        pdf.set_font("Arial", "", 8)
        for sex, count in sorted(sex_counts.items()):
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            pdf.cell(40, 8, sex, border=1, align="L")
            pdf.cell(30, 8, str(count), border=1, align="C")
            pdf.cell(30, 8, f"{percentage:.1f}%", border=1, align="C")
            pdf.ln()
        
        pdf.ln(5)
        
        # Barangay Distribution Table
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, "Barangay Distribution", ln=1)

        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(60, 8, "Barangay", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Count", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Percentage", border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Arial", "", 8)
        # Sort barangays: count descending, then alphabetically for ties
        all_barangays = sorted(barangay_counts.items(), key=lambda x: (-x[1], x[0].upper()))

        for barangay, count in all_barangays:
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            barangay_name = barangay[:25]
            pdf.cell(60, 8, barangay_name, border=1, align="L")
            pdf.cell(30, 8, str(count), border=1, align="C")
            pdf.cell(30, 8, f"{percentage:.1f}%", border=1, align="C")
            pdf.ln()

        pdf.ln(5)

    else:
        pdf.set_font("Arial", "I", 10)
        pdf.cell(0, 8, "No patients were flagged by the AI during this period.", ln=1)
    
    pdf.ln(5)

    # Review Status Confirmation
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 9, "Review Status:", ln=1)
    pdf.set_font("Arial", "", 10)
    pending_count = data.get("Pending Confirmations", "0")
    if pending_count == "0":
        status_text = "Zero (0) patients are currently pending clinical confirmation, ensuring all flagged cases have been resolved."
    else:
        status_text = f"{pending_count} patient(s) are currently pending clinical confirmation and require review."
    pdf.multi_cell(0, 6, status_text, align="J")
    pdf.ln(5)

    # Disclaimer Section
    disclaimer_text = (
        "This report reflects preliminary results generated by the DeteXTB system, "
        "intended for informational purposes only and should not be used to make clinical decisions. "
        "All flagged cases must undergo further review by a licensed medical professional."
    )

    pdf.set_text_color(100, 100, 100) 
    pdf.set_font("Arial", "I", 10) 
    pdf.multi_cell(0, 6, disclaimer_text, align="J")
    pdf.ln(15)

    # Signature Line + Text
    pdf.set_text_color(0, 0, 0)  
    line_y = pdf.get_y()  
    pdf.line(70, line_y, 140, line_y)  
    pdf.ln(2) 
    pdf.set_font("Arial", "I", 11)
    pdf.cell(0, 10, "Signature", ln=1, align="C")

    pdf.ln(5)

    # Timestamp
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 10, f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1, align='R')

    # Reset text color to black
    pdf.set_text_color(0, 0, 0)

    return bytes(pdf.output(dest='S'))

# Function to generate Confirmed TB Cases Report PDF
def generate_confirmed_pdf(report_title, data, filter_info=None, selected_month=None, selected_year=None):
    pdf = PDFReport_format()
    pdf.add_page()

    # Report title
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, report_title, ln=1, align="C")
    pdf.ln(5)

    # Add filter information if provided
    if filter_info:
        pdf.set_font("Arial", "I", 10)
        pdf.cell(0, 8, filter_info, ln=1, align="C")
        pdf.ln(5)

    pdf.set_draw_color(0, 0, 0)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    # Colors
    red = (217, 57, 53)
    green = (0, 128, 0)
    black = (0, 0, 0)

    # Force black text initially
    pdf.set_text_color(*black)

    # Compute percentages for display
    total_text = data.get("Total Confirmed Cases", "0 Confirmed TB Cases")
    total_count = int(total_text.split()[0])

    positive_text = data.get("Positive Cases", "0")
    negative_text = data.get("Negative Cases", "0")

    positive_count = int(positive_text.split()[0])
    negative_count = int(negative_text.split()[0])

    # Avoid division by zero
    pos_pct = (positive_count / total_count * 100) if total_count > 0 else 0
    neg_pct = (negative_count / total_count * 100) if total_count > 0 else 0

    # Update text to include percentage 
    total_text = f"{total_count}"
    positive_text = f"{positive_count} ({pos_pct:.0f}% of total)"
    negative_text = f"{negative_count} ({neg_pct:.0f}% of total)"

    # Replace the data entries for consistency in the summary section
    data["Positive Cases"] = positive_text
    data["Negative Cases"] = negative_text
    data["Total Confirmed Cases"] = total_text

    # Summary section
    for key, value in data.items():
        # No italicization for percentages anymore
        italicize = False  

        # Custom coloring for specific fields
        if key.lower().startswith("positive"):
            color = red
        elif key.lower().startswith("negative"):
            color = green
        else:
            color = black

        pdf.write_key_value(key, value, value_color=color, italic_in_parens=italicize)

    pdf.ln(5)

    # Visual Summary Section (Pie Chart)
    pdf.set_font("Arial", "B", 12)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 10, "Visual Summary", ln=1)

    if total_count > 0:
        fig, ax = plt.subplots(figsize=(6.5, 6.5))
        labels = ['Positive', 'Negative']
        sizes = [positive_count, negative_count]
        colors = ['#E53935', "#5FCA62"]
        explode = (0.05, 0.05)

        wedges, texts, autotexts = ax.pie(
            sizes, explode=explode, labels=labels, colors=colors,
            autopct='%1.0f%%', shadow=True, startangle=90,
            textprops={'fontsize': 18}
        )

        for t in texts:
            t.set_fontsize(18)

        for t in autotexts:
            t.set_fontsize(16)

        ax.axis('equal')

        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=200)
        img_buffer.seek(0)
        plt.close()
        pdf.image(img_buffer, x=55, w=100)
        pdf.ln(10)
    else:
        pdf.set_font("Arial", "I", 11)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 10, "No data available for visual summary.", ln=1, align="C")
        pdf.ln(10)

    # Confirmed Cases Details Section
    case_details = fetch_confirmed_case_details(selected_month, selected_year)
    if case_details:
        try:
            case_details.sort(
                key=lambda x: (
                    datetime.strptime(str(x.get("Confirmation Date", ""))[:10], "%Y-%m-%d")
                    if x.get("Confirmation Date") else datetime.min,
                    str(x.get("Patient ID", "")).lower()
                ),
                reverse=True
            )
        except Exception as e:
            print("Sorting error (Confirmed Cases):", e)

    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Confirmed Cases Details", ln=1)

    if case_details:
        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(200, 200, 200)

        col_widths = [8, 27, 8, 8, 23, 35, 30, 33, 18] 
        headers = ["#", "Patient ID", "Age", "Sex", "Barangay", "Final Status", "Confirmation Date", "Confirmation Method", "AI Flagged"]

        total_table_width = sum(col_widths)
        if total_table_width > 190:
            scale_factor = 190 / total_table_width
            col_widths = [int(w * scale_factor) for w in col_widths]

        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 8, header, border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Arial", "", 7) 
        for idx, case in enumerate(case_details, start=1):
            status = str(case["Final Status"]).lower()
            ai_flagged = str(case["AI Flagged"]).lower()

            if "positive" in status and "negative" in ai_flagged:
                pdf.set_fill_color(200, 150, 255)
                fill = True
            elif "negative" in status and "positive" in ai_flagged:
                pdf.set_fill_color(255, 200, 100)
                fill = True
            elif "positive" in status:
                pdf.set_fill_color(255, 150, 150)
                fill = True
            elif "negative" in status:
                pdf.set_fill_color(150, 255, 150)
                fill = True
            else:
                fill = False

            patient_id = str(case["Patient ID"])
            barangay = str(case["Barangay"])
            final_status = str(case["Final Status"])
            conf_method = str(case["Confirmation Method"])

            pdf.cell(col_widths[0], 8, str(idx), border=1, align="C", fill=fill)
            pdf.cell(col_widths[1], 8, patient_id, border=1, align="C", fill=fill)
            pdf.cell(col_widths[2], 8, str(case["Age"]), border=1, align="C", fill=fill)
            pdf.cell(col_widths[3], 8, str(case["Sex"]), border=1, align="C", fill=fill)
            pdf.cell(col_widths[4], 8, barangay, border=1, align="C", fill=fill)
            pdf.cell(col_widths[5], 8, final_status, border=1, align="C", fill=fill)
            pdf.cell(col_widths[6], 8, str(case["Confirmation Date"])[:10], border=1, align="C", fill=fill)
            pdf.cell(col_widths[7], 8, conf_method, border=1, align="C", fill=fill)
            pdf.cell(col_widths[8], 8, str(case["AI Flagged"]), border=1, align="C", fill=fill)
            pdf.ln()

        # Demographic Summary Section
        pdf.ln(8)

        # Calculate demographic summaries
        age_groups = {
            "Children (0-14)": 0,
            "Youth/Young Adults (15-24)": 0,
            "Adults (25-64)": 0,
            "Elderly (65+)": 0
        }

        sex_counts = {"Female": 0, "Male": 0, "Unknown": 0}
        barangay_counts = {}

        for case in case_details:
            # Age group calculation
            age = case.get("Age")
            if age and str(age).isdigit():
                age_int = int(age)
                if age_int <= 14:
                    age_groups["Children (0-14)"] += 1
                elif age_int <= 24:
                    age_groups["Youth/Young Adults (15-24)"] += 1
                elif age_int <= 64:
                    age_groups["Adults (25-64)"] += 1
                else:
                    age_groups["Elderly (65+)"] += 1

            # Sex count
            raw_sex = str(case.get("Sex", "Unknown")).strip().upper()
            if raw_sex == "F":
                sex = "Female"
            elif raw_sex == "M":
                sex = "Male"
            else:
                sex = "Unknown"
            sex_counts[sex] += 1

            # Barangay count
            barangay = case.get("Barangay", "Unknown")
            barangay_counts[barangay] = barangay_counts.get(barangay, 0) + 1

        total_patients = len(case_details)

        all_barangays = sorted(barangay_counts.items(), key=lambda x: x[1], reverse=True) # descending

        # Age Group Distribution Table 
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, "Age Group Distribution", ln=1)

        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(60, 8, "Age Group", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Count", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Percentage", border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Arial", "", 8)
        for age_group, count in age_groups.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            pdf.cell(60, 8, age_group, border=1, align="L")
            pdf.cell(30, 8, str(count), border=1, align="C")
            pdf.cell(30, 8, f"{percentage:.1f}%", border=1, align="C")
            pdf.ln()
        pdf.ln(5)

        # Sex Distribution Table
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, "Sex Distribution", ln=1)

        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(40, 8, "Sex", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Count", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Percentage", border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Arial", "", 8)
        for sex in ["Female", "Male", "Unknown"]:
            count = sex_counts.get(sex, 0)
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            pdf.cell(40, 8, sex, border=1, align="L")
            pdf.cell(30, 8, str(count), border=1, align="C")
            pdf.cell(30, 8, f"{percentage:.1f}%", border=1, align="C")
            pdf.ln()
        pdf.ln(5)

        # Barangay Distribution Table
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, "Barangay Distribution", ln=1)

        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(60, 8, "Barangay", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Count", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Percentage", border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Arial", "", 8)
        # Sort barangays: first by count descending, then alphabetically for ties
        all_barangays = sorted(barangay_counts.items(), key=lambda x: (-x[1], x[0].upper()))

        for barangay, count in all_barangays:
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            barangay_name = barangay[:25]
            pdf.cell(60, 8, barangay_name, border=1, align="L")
            pdf.cell(30, 8, str(count), border=1, align="C")
            pdf.cell(30, 8, f"{percentage:.1f}%", border=1, align="C")
            pdf.ln()
        pdf.ln(5)

    else:
        pdf.set_font("Arial", "I", 10)
        pdf.cell(0, 8, "No confirmed cases during this period.", ln=1)

    pdf.ln(10)

    # Disclaimer Section
    disclaimer_text = (
        "This report summarizes confirmed TB cases recorded in the DeteXTB system, "
        "intended for record-keeping and statistical purposes only. "
        "Clinical decisions should always be based on professional medical evaluation."
    )
    pdf.set_text_color(100, 100, 100)
    pdf.set_font("Arial", "I", 10)
    pdf.multi_cell(0, 6, disclaimer_text, align="J")
    pdf.ln(15)

    # Signature Line + Text
    pdf.set_text_color(0, 0, 0)
    line_y = pdf.get_y()
    pdf.line(70, line_y, 140, line_y)
    pdf.ln(2)
    pdf.set_font("Arial", "I", 11)
    pdf.cell(0, 10, "Signature", ln=1, align="C")
    pdf.ln(5)

    # Timestamp
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 10, f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1, align='R')
    pdf.set_text_color(0, 0, 0)

    return bytes(pdf.output(dest='S'))

# Function to generate AI Presumptive TB Report Excel
def generate_ai_excel(report_title, data, filter_info=None, selected_month=None, selected_year=None):
    wb = Workbook()
    
    # --- Worksheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'] = "Mandaue City Health Office"
    ws_summary['A1'].font = Font(bold=True, size=14)

    ws_summary.merge_cells('A2:B2')
    ws_summary['A2'] = "S.B. Cabahug, Mandaue City, Philippines."

    ws_summary.merge_cells('A3:B3')
    ws_summary['A3'] = "Call us on: +63 (032) 230 4500 | FB: Mandaue City Public Affairs Office | Email: cmo@mandauecity.gov.ph"

    ws_summary.merge_cells('A5:B5')
    ws_summary['A5'] = "DeteXTB: AI-Assisted Presumptive Tuberculosis Detection and Mapping System"
    ws_summary['A5'].font = Font(bold=True, size=12)

    ws_summary.merge_cells('A7:B7')
    ws_summary['A7'] = report_title
    ws_summary['A7'].font = Font(bold=True, size=16)

    row_start = 9
    if filter_info:
        ws_summary.merge_cells(f'A{row_start}:B{row_start}')
        ws_summary[f'A{row_start}'] = filter_info
        ws_summary[f'A{row_start}'].font = Font(italic=True)
        row_start += 2

    ws_summary[f'A{row_start}'] = "Summary"
    ws_summary[f'A{row_start}'].font = Font(bold=True)

    colors = {"Pending": "FFE699", "Positive": "FF9999", "Negative": "99FF99"}

    row = row_start + 1
    for key, value in data.items():
        ws_summary[f"A{row}"] = key
        ws_summary[f"B{row}"] = value
        ws_summary[f"B{row}"].alignment = Alignment(horizontal="left")
        for status, color_code in colors.items():
            if status in key:
                fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                ws_summary[f"A{row}"].fill = fill
                ws_summary[f"B{row}"].fill = fill
                break
        row += 1

    # Disclaimer
    disclaimer_text = (
        "This report reflects preliminary results generated by the DeteXTB system, "
        "intended for informational purposes only and should not be used to make clinical decisions. "
        "All flagged cases must undergo further review by a licensed medical professional."
    )

    import textwrap
    wrapped_lines = textwrap.wrap(disclaimer_text, width=100)
    for i, line in enumerate(wrapped_lines):
        curr_row = row + 2 + i
        ws_summary.merge_cells(f'A{curr_row}:B{curr_row}')
        ws_summary[f'A{curr_row}'] = line
        ws_summary[f'A{curr_row}'].font = Font(italic=True, color="808080")

    timestamp_row = row + 2 + len(wrapped_lines) + 2
    ws_summary.merge_cells(f'A{timestamp_row}:B{timestamp_row}')
    ws_summary[f"A{timestamp_row}"] = f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary[f"A{timestamp_row}"].font = Font(italic=True)

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 65

    # --- Worksheet 2: Flagged Patients Log ---
    ws_patients = wb.create_sheet(title="Flagged Patients Log")
    ws_patients['A1'] = "Flagged Patients Log"
    ws_patients['A1'].font = Font(bold=True, size=14)
    ws_patients.merge_cells('A1:K1')

    if filter_info:
        ws_patients['A2'] = filter_info
        ws_patients['A2'].font = Font(italic=True)
        ws_patients.merge_cells('A2:K2')
        header_row = 4
    else:
        header_row = 3

    headers = ["#", "Patient ID", "Age", "Sex", "Barangay", "", "AI Flagged Date", "AI Confidence Level", "", "Final Status", "Confirmation Date"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_patients.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if header.strip() != "":
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Fetch patient data
    patient_details = fetch_flagged_patient_details(selected_month, selected_year)
    try:
        patient_details.sort(
            key=lambda x: (
                datetime.strptime(str(x.get("AI Flagged Date", ""))[:10], "%Y-%m-%d")
                if x.get("AI Flagged Date") else datetime.max,
                str(x.get("Patient ID", "")).lower()
            )
        )
    except Exception as e:
        print("Sorting error (AI Excel):", e)

    # Data rows
    data_row = header_row + 1
    for i, patient in enumerate(patient_details, start=1):
        ai_date = patient.get("AI Flagged Date")
        if ai_date:
            try: ai_date = datetime.fromisoformat(str(ai_date)).strftime("%Y-%m-%d")
            except: pass

        confidence = patient.get("AI Confidence Score")
        if confidence is not None:
            try: confidence = f"{round(float(confidence) * 100):d}%"
            except: pass

        confirmation_date = patient.get("Confirmation Date") or "Not yet confirmed"

        row_values = [
            i, patient["Patient ID"], patient["Age"], patient["Sex"], patient["Barangay"],
            None, ai_date, confidence, None, patient["Final Status"], confirmation_date
        ]
        fill_color = ("FF9999" if "Positive" in patient["Final Status"] else
                      "99FF99" if "Negative" in patient["Final Status"] else "FFE699")

        for col_idx, value in enumerate(row_values, start=1):
            ws_patients.cell(row=data_row, column=col_idx, value=value)
            ws_patients.cell(row=data_row, column=col_idx).alignment = Alignment(horizontal="center")
            if value is not None:
                ws_patients.cell(row=data_row, column=col_idx).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
        data_row += 1

    # Column widths
    widths = [6, 30, 8, 8, 20, 3, 15, 18, 3, 20, 18]
    for idx, width in enumerate(widths, start=1):
        ws_patients.column_dimensions[chr(64 + idx)].width = width

    # --- Worksheet 3: Flagged Patients Demographics ---
    if patient_details:
        age_groups = {"Children (0-14)": 0, "Youth/Young Adults (15-24)": 0, "Adults (25-64)": 0, "Elderly (65+)": 0}
        sex_counts = {}
        barangay_counts = {}

        # Populate counts
        for patient in patient_details:
            age = patient.get("Age")
            if age and str(age).isdigit():
                a = int(age)
                if a <= 14: age_groups["Children (0-14)"] += 1
                elif a <= 24: age_groups["Youth/Young Adults (15-24)"] += 1
                elif a <= 64: age_groups["Adults (25-64)"] += 1
                else: age_groups["Elderly (65+)"] += 1

            raw_sex = patient.get("Sex", "Unknown").strip().upper()
            sex = "Female" if raw_sex == "F" else "Male" if raw_sex == "M" else "Unknown"
            sex_counts[sex] = sex_counts.get(sex, 0) + 1

            barangay = patient.get("Barangay", "Unknown")
            barangay_counts[barangay] = barangay_counts.get(barangay, 0) + 1

        total_patients = len(patient_details)
        ws_demo = wb.create_sheet(title="Flagged Patients Demographics")

        ws_demo['A1'] = "Flagged Patients Demographics"
        ws_demo['A1'].font = Font(bold=True, size=14)
        ws_demo.merge_cells('A1:K1')

        if filter_info:
            ws_demo['A2'] = filter_info
            ws_demo['A2'].font = Font(italic=True)
            ws_demo.merge_cells('A2:K2')

        starting_row = 4

        # Function to write horizontal tables
        def write_demo_table_horizontal(ws, title, data_dict, start_col, total, start_row):
            ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True, size=12)
            headers = ["Category", "Count", "Percentage"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=start_row+1, column=start_col+i, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            row = start_row + 2
            for key, count in data_dict.items():
                ws.cell(row=row, column=start_col, value=key)
                ws.cell(row=row, column=start_col+1, value=count).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=start_col+2, value=f"{(count/total*100):.1f}%").alignment = Alignment(horizontal="center")
                row += 1

            # Column widths
            for col_idx in range(start_col, start_col + len(headers)):
                max_length = 0
                for r in range(start_row+1, row):
                    cell_val = ws.cell(row=r, column=col_idx).value
                    if cell_val is not None:
                        max_length = max(max_length, len(str(cell_val)))
                ws.column_dimensions[chr(64 + col_idx)].width = max_length + 5

            return start_col + len(headers) + 1

        # Write tables horizontally
        col = 1
        col = write_demo_table_horizontal(ws_demo, "Age Group Distribution", age_groups, col, total_patients, starting_row)
        sex_ordered = {k: sex_counts.get(k, 0) for k in ["Female", "Male", "Unknown"]}
        col = write_demo_table_horizontal(ws_demo, "Sex Distribution", sex_ordered, col, total_patients, starting_row)
        sorted_barangays = dict(sorted(barangay_counts.items(), key=lambda x: (-x[1], x[0].upper())))
        col = write_demo_table_horizontal(ws_demo, "Barangay Distribution", sorted_barangays, col, total_patients, starting_row)

    # --- Worksheet 4: AI Performance Metrics ---
    ws_metrics = wb.create_sheet(title="AI Performance Metrics")
    
    ws_metrics['A1'] = "AI Performance Metrics"
    ws_metrics['A1'].font = Font(bold=True, size=14)
    ws_metrics.merge_cells('A1:B1')
    
    if filter_info:
        ws_metrics['A2'] = filter_info
        ws_metrics['A2'].font = Font(italic=True)
        ws_metrics.merge_cells('A2:B2')
        metrics_start_row = 4
    else:
        metrics_start_row = 3
    
    # Performance metrics data
    metrics = calculate_ai_performance_metrics(selected_month, selected_year)
    
    ws_metrics[f'A{metrics_start_row}'] = "Metric"
    ws_metrics[f'B{metrics_start_row}'] = "Count"
    ws_metrics[f'A{metrics_start_row}'].font = Font(bold=True)
    ws_metrics[f'B{metrics_start_row}'].font = Font(bold=True)
    ws_metrics[f'A{metrics_start_row}'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws_metrics[f'B{metrics_start_row}'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    metrics_row = metrics_start_row + 1
    for metric_name, metric_value in metrics.items():
        ws_metrics[f'A{metrics_row}'] = metric_name
        ws_metrics[f'B{metrics_row}'] = metric_value
        ws_metrics[f'B{metrics_row}'].alignment = Alignment(horizontal="left")
        metrics_row += 1
    
    ws_metrics.column_dimensions['A'].width = 30
    ws_metrics.column_dimensions['B'].width = 10

    # Save to BytesIO
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()

# Function to generate Confirmed TB Cases Report Excel
def generate_confirmed_excel(report_title, data, filter_info=None, selected_month=None, selected_year=None):
    wb = Workbook()

    # --- Worksheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Header
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = "Mandaue City Health Office"
    ws_summary['A1'].font = Font(bold=True, size=14)

    ws_summary.merge_cells('A2:D2')
    ws_summary['A2'] = "S.B. Cabahug, Mandaue City, Philippines."

    ws_summary.merge_cells('A3:D3')
    ws_summary['A3'] = "Call us on: +63 (032) 230 4500 | FB: Mandaue City Public Affairs Office | Email: cmo@mandauecity.gov.ph"

    ws_summary.merge_cells('A5:D5')
    ws_summary['A5'] = "DeteXTB: AI-Assisted Presumptive Tuberculosis Detection and Mapping System"
    ws_summary['A5'].font = Font(bold=True, size=12)

    ws_summary.merge_cells('A7:D7')
    ws_summary['A7'] = report_title
    ws_summary['A7'].font = Font(bold=True, size=16)

    row_start = 9
    if filter_info:
        ws_summary.merge_cells(f'A{row_start}:D{row_start}')
        ws_summary[f'A{row_start}'] = filter_info
        ws_summary[f'A{row_start}'].font = Font(italic=True)
        row_start += 2

    # Summary table headers
    headers = ["Total Confirmed", "Total Positive", "Total Negative"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=row_start, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Extract counts
    import re

    total_text = data.get("Total Confirmed Cases", "0 Confirmed TB Cases")
    positive_text = data.get("Positive Cases", "0 0% of total")
    negative_text = data.get("Negative Cases", "0 0% of total")

    total_count = int(total_text.split()[0])
    positive_count = int(positive_text.split()[0])
    negative_count = int(negative_text.split()[0])

    # Extract percentages using regex
    pos_match = re.search(r'(\d+)%', positive_text)
    neg_match = re.search(r'(\d+)%', negative_text)
    pos_percent = pos_match.group(1) + "%" if pos_match else "0%"
    neg_percent = neg_match.group(1) + "%" if neg_match else "0%"

    # Combine count and percentage in one string
    positive_display = f"{positive_count} ({pos_percent})"
    negative_display = f"{negative_count} ({neg_percent})"

    data_row = row_start + 1
    ws_summary[f'A{data_row}'] = total_count
    ws_summary[f'B{data_row}'] = positive_display
    ws_summary[f'C{data_row}'] = negative_display

    # Color fill
    if positive_count > 0:
        ws_summary[f'B{data_row}'].fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    if negative_count > 0:
        ws_summary[f'C{data_row}'].fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    for col in ['A', 'B', 'C', 'D']:
        ws_summary[f'{col}{data_row}'].alignment = Alignment(horizontal="center")

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 30

    # Disclaimer
    disclaimer_lines = [
        "This report summarizes confirmed TB cases recorded in the DeteXTB system, intended for record-keeping",
        "and statistical purposes only. Clinical decisions should always be based on professional medical evaluation."
    ]

    for i, line in enumerate(disclaimer_lines):
        curr_row = data_row + 3 + i
        ws_summary.merge_cells(f'A{curr_row}:D{curr_row}')
        ws_summary[f'A{curr_row}'] = line
        ws_summary[f'A{curr_row}'].font = Font(italic=True, color="808080")
        ws_summary[f'A{curr_row}'].alignment = Alignment(horizontal="left", wrap_text=True)

    timestamp_row = data_row + 3 + len(disclaimer_lines) + 2
    ws_summary.merge_cells(f'A{timestamp_row}:D{timestamp_row}')
    ws_summary[f"A{timestamp_row}"] = f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary[f"A{timestamp_row}"].font = Font(italic=True)

    # --- Worksheet 2: Confirmed Cases Details ---
    ws_register = wb.create_sheet(title="Confirmed Cases Details")
    ws_register['A1'] = "Confirmed Cases Details"
    ws_register['A1'].font = Font(bold=True, size=14)
    ws_register.merge_cells('A1:L1')

    if filter_info:
        ws_register['A2'] = filter_info
        ws_register['A2'].font = Font(italic=True)
        ws_register.merge_cells('A2:L2')
        header_row = 4
    else:
        header_row = 3

    headers = ["#", "Patient ID", "Age", "Sex", "Barangay", "", "Final Status", "Confirmation Date", "Confirmation Method", "", "AI Flagged"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_register.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if header.strip() != "":
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Fetch and sort case details
    case_details = fetch_confirmed_case_details(selected_month, selected_year) or []

    try:
        case_details.sort(
            key=lambda x: (
                datetime.strptime(str(x.get("Confirmation Date", ""))[:10], "%Y-%m-%d")
                if x.get("Confirmation Date") else datetime.min,
                str(x.get("Patient ID", "")).lower()
            ),
            reverse=True
        )
    except Exception as e:
        print("Sorting error (Excel Confirmed Cases):", e)

    # Write main table data
    data_row = header_row + 1
    for i, case in enumerate(case_details, start=1):
        confirmation_date = case.get("Confirmation Date", "")
        if "T" in confirmation_date:
            confirmation_date = confirmation_date.split("T")[0]

        # None is spacer
        row_values = [
            i, case.get("Patient ID"), case.get("Age"), case.get("Sex"), case.get("Barangay"), None, 
            case.get("Final Status"), confirmation_date, 
            case.get("Confirmation Method"), None, 
            case.get("AI Flagged")
        ]

        status = str(case.get("Final Status", "")).lower()
        ai_flagged = str(case.get("AI Flagged", "")).lower()
        if "positive" in status and "negative" in ai_flagged:
            fill_color = "C896FF"
        elif "negative" in status and "positive" in ai_flagged:
            fill_color = "FFD580"
        elif "positive" in status:
            fill_color = "FF9999"
        elif "negative" in status:
            fill_color = "99FF99"
        else:
            fill_color = "FFFFFF"

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws_register.cell(row=data_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center")
            if value is not None:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        data_row += 1

    column_widths = [6, 30, 8, 8, 20, 3, 22, 20, 22, 3, 18]
    for col_idx, width in enumerate(column_widths, start=1):
        ws_register.column_dimensions[chr(64 + col_idx)].width = width

    # --- Worksheet 3: Confirmed Cases Demographics ---
    if case_details:
        age_groups = {"Children (0-14)": 0, "Youth/Young Adults (15-24)": 0, "Adults (25-64)": 0, "Elderly (65+)": 0}
        sex_counts = {}
        barangay_counts = {}

        for case in case_details:
            age = case.get("Age")
            if age and str(age).isdigit():
                age_int = int(age)
                if age_int <= 14: age_groups["Children (0-14)"] += 1
                elif age_int <= 24: age_groups["Youth/Young Adults (15-24)"] += 1
                elif age_int <= 64: age_groups["Adults (25-64)"] += 1
                else: age_groups["Elderly (65+)"] += 1

            raw_sex = case.get("Sex", "Unknown").strip().upper()
            sex = "Female" if raw_sex == "F" else "Male" if raw_sex == "M" else "Unknown"
            sex_counts[sex] = sex_counts.get(sex, 0) + 1

            barangay = case.get("Barangay", "Unknown")
            barangay_counts[barangay] = barangay_counts.get(barangay, 0) + 1

        total_cases = len(case_details)
        ws_demo = wb.create_sheet(title="Confirmed Cases Demographics")
        ws_demo['A1'] = "Confirmed Cases Demographics"
        ws_demo['A1'].font = Font(bold=True, size=14)
        ws_demo.merge_cells('A1:L1')

        if filter_info:
            ws_demo['A2'] = filter_info
            ws_demo['A2'].font = Font(italic=True)
            ws_demo.merge_cells('A2:L2')

        starting_row = 4 

        def write_demo_table_horizontal(ws, title, data_dict, start_col, total, start_row):
            ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True, size=12)
            headers = ["Category", "Count", "Percentage"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=start_row+1, column=start_col+i, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            row = start_row + 2
            for key, count in data_dict.items():
                ws.cell(row=row, column=start_col, value=key)
                ws.cell(row=row, column=start_col+1, value=count).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=start_col+2, value=f"{(count/total*100):.1f}%").alignment = Alignment(horizontal="center")
                row += 1

            for col_idx in range(start_col, start_col + len(headers)):
                max_length = 0
                for r in range(start_row+1, row):
                    cell_val = ws.cell(row=r, column=col_idx).value
                    if cell_val is not None:
                        max_length = max(max_length, len(str(cell_val)))
                ws.column_dimensions[chr(64 + col_idx)].width = max_length + 5

            return start_col + len(headers) + 1

        # Write tables horizontally
        col = 1
        col = write_demo_table_horizontal(ws_demo, "Age Group Distribution", age_groups, col, total_cases, starting_row)
        sex_ordered = {k: sex_counts.get(k, 0) for k in ["Female", "Male", "Unknown"]}
        col = write_demo_table_horizontal(ws_demo, "Sex Distribution", sex_ordered, col, total_cases, starting_row)
        sorted_barangays = dict(sorted(barangay_counts.items(), key=lambda x: (-x[1], x[0].upper())))
        col = write_demo_table_horizontal(ws_demo, "Barangay Distribution", sorted_barangays, col, total_cases, starting_row)

    # Save to BytesIO
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


# Function to format filter information for reports
def format_filter_info(selected_month, selected_year):
    if selected_month == "All":
        month_text = "Year"
    else:
        month_text = datetime(2000, selected_month, 1).strftime('%B')
    
    return f"Reporting Period: {month_text} {selected_year}"


# --- Main UI ---

def Reports(is_light=True):
    # Initialize theme state in session if not set
    if "light_mode" not in st.session_state:
        st.session_state["light_mode"] = True

    if is_light is None:
        is_light = st.session_state["light_mode"]

    if "report_filters_presumptive" not in st.session_state:
        current_month = datetime.today().month
        current_year = datetime.today().year
        
        st.session_state["report_filters_presumptive"] = {
            "selected_month_presumptive": current_month,
            "selected_year_presumptive": current_year,
        }

    if "report_filters_confirmed" not in st.session_state:
        current_month = datetime.today().month
        current_year = datetime.today().year
        
        st.session_state["report_filters_confirmed"] = {
            "selected_month_confirmed": current_month,
            "selected_year_confirmed": current_year,
        }

    # Theme toggle UI element
    col_title, col_toggle = st.columns([6, 1])
    with col_title:
        st.markdown("<h4>Reports</h4>", unsafe_allow_html=True)
    with col_toggle:
        new_toggle = st.toggle("🌙", value=is_light, key="theme_toggle", label_visibility="collapsed")

    # Rerun app if theme toggle changes
    if new_toggle != st.session_state["light_mode"]:
        st.session_state["light_mode"] = new_toggle
        st.rerun()

    # Define colors based on theme
    is_light = st.session_state["light_mode"]
    bg_color = "white" if is_light else "#0e0e0e"
    text_color = "black" if is_light else "white"
    card_color = "#D9D9D9" if is_light else "#1e1e1e"
    card_text_color = "black" if is_light else "white"
    title_color = "black" if is_light else "white"
    button_bg = "#d32f2f"
    button_hover = "#f3a5a5"
    input_bg = "white" if is_light else "#1e1e1e"
    input_border = "black" if is_light else "white"
    placeholder_color = "#b0b0b0" if is_light else "#888"
    calendar_dropdown_bg = "white" if is_light else "#222"
    calendar_border = "black" if is_light else "none"
    select_bg = "white" if is_light else "#1e1e1e"
    select_hover = "#f0f0f0" if is_light else "#2a2a2a"

    # Inject CSS styles for the app UI
    st.markdown(f"""
    <style>
        .stApp {{
            padding-top: 0rem !important;
        }}
        .block-container {{
            padding-top: 2rem !important;
            padding-left: 2rem !important;
            padding-right: 2rem !important;
        }}
        h1, h2, h3, h4, h5, h6 {{
            color: {title_color} !important;
        }}
        html, body, [class*="css"] {{
            color: {text_color} !important;
            font-family: 'Arial', sans-serif;
        }}
        [data-testid="stAppViewContainer"], [data-testid="stAppViewContainer"] > .main {{
            background-color: {bg_color} !important;
            color: {text_color} !important;
        }}
        /* Text input */
        .stTextInput > div {{
            background-color: {input_bg} !important;
            border-radius: 25px !important;
            padding: 2px !important;
            border: 1px solid {input_border} !important;
        }}
        .stTextInput input {{
            background-color: {input_bg} !important;
            border: none !important;
            color: {text_color} !important;
            padding-left: 10px !important;
            font-size: 1rem !important;
            caret-color: {"black" if is_light else "white"} !important;
        }}
        .stTextInput input::placeholder {{
            color: {placeholder_color} !important;
        }}

        div.stTextInput > div > div > div[data-testid="InputInstructions"],
        [data-testid="InputInstructions"] {{
            display: none !important;
        }}

        /* Date input */
        div[data-testid="stDateInput"] > div {{
            background-color: {input_bg} !important;
            border-radius: 25px !important;
            border: 1px solid {input_border} !important;
            padding: 0 10px !important;
        }}
        div[data-testid="stDateInput"] > div > div {{
            background-color: transparent !important;
            border: none !important;
            box-shadow: none !important;
        }}
        div[data-testid="stDateInput"] input {{
            background-color: {input_bg} !important;
            border: none !important;
            color: {text_color} !important;
            padding: 8px 0 !important;
            box-shadow: none !important;
        }}
        div[data-testid="stDateInput"] svg {{
            color: {text_color} !important;
        }}

        div[data-baseweb="popover"] {{
            background-color: {calendar_dropdown_bg} !important;
            border-radius: 15px !important;
            border: 1px solid {calendar_border} !important;
            box-shadow: none !important;
        }}
        
        div[data-testid="stDateInput"] > div:focus-within {{
            box-shadow: none !important;
        }}

        /* Selectboxes */
        .stSelectbox div[data-baseweb="select"] > div,
        .stSelectbox div[role="listbox"],
        .stSelectbox li {{
            margin-bottom: 20px !important;
            background-color: {select_bg} !important;
            color: {text_color} !important;
            border: 1px solid {input_border} !important;
            border-radius: 25px !important;
        }}
        .stSelectbox li:hover {{
            background-color: {select_hover} !important;
        }}
        .stSelectbox label p {{
            color: {title_color} !important;  
            font-weight: bold !important;    
        }}
        .card-title {{
            font-size: 23px;
            font-weight: bold;
            color: {title_color};
            margin-bottom: 15px;
            margin-top: 30px;
        }}
        .card-container {{
            display: flex;
            gap: 20px;
            margin-top: 5px;
        }}
        .report-card {{
            background-color: {card_color};
            border-radius: 12px;
            padding: 20px;
            flex: 1;
            text-align: center;
            color: {card_text_color};
        }}
        .report-card .line-1 {{
            font-weight: bold;
            font-size: 20px;
            margin-bottom: 5px;
        }}
        .report-card .line-2 {{
            font-size: 35px;
            color: #E53935;
            font-weight: 600;
            margin-bottom: 5px;
        }}
        .report-card .line-3 {{
            font-size: 20px;
            font-weight: normal;
        }}
        div.stDownloadButton > button {{
            background-color: {button_bg} !important;
            color: white !important;
            font-weight: bold !important;
            border: none !important;
            border-radius: 25px !important;
            padding: 8px 20px !important;
            cursor: pointer !important;
            font-size: 1rem !important;
            min-width: 120px !important;
            white-space: nowrap !important;
            margin-left: 8px !important;
            transition: background-color 0.3s ease !important;
        }}
        div.stDownloadButton > button:hover {{
            background-color: {button_hover} !important;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.2) !important;
        }}
        /* Custom button styling ONLY for main content, NOT sidebar */
        .block-container div[data-testid="stButton"] > button,
        .block-container div[data-testid^="FormSubmitter"] button {{
            background-color: {button_bg} !important;
            color: white !important;
            font-weight: bold !important;
            border: none !important;
            border-radius: 25px !important;
            padding: 10px 20px !important;
            font-size: 1rem !important;
            margin-left: -5px;
            margin-top: 6px;
            min-width: 120px !important;
            max-width: 180px !important;
            white-space: nowrap !important;
        }}

        .block-container div[data-testid="stButton"] > button:hover,
        .block-container div[data-testid^="FormSubmitter"] button:hover {{
            background-color: {button_hover} !important;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.2) !important;
        }}
        .filter-label {{
            color: {title_color};
            margin-bottom: 5px;
        }}
    </style>
    """, unsafe_allow_html=True)


    # --- AI Presumptive TB Report Block ---

    # Display AI report title + filters + reset button (under filters)
    col_title, col_filters = st.columns([6, 6])
    with col_title:
        st.markdown('<div class="card-title">AI Presumptive TB Report</div>', unsafe_allow_html=True)

    with col_filters:
        col_month, col_year, col_reset = st.columns([1,1,0.6])

        # Initialize/reset counter
        if "presumptive_reset_count" not in st.session_state:
            st.session_state["presumptive_reset_count"] = 0
        reset_count = st.session_state["presumptive_reset_count"]

        # Month
        with col_month:
            st.markdown('<div class="filter-label">Month</div>', unsafe_allow_html=True)
            current_stored = st.session_state["report_filters_presumptive"].get(
                "selected_month_presumptive", datetime.today().month
            )
            month_options = ["All"] + list(range(1, 13))
            default_index = 0 if current_stored == "All" else month_options.index(current_stored)
            new_month = st.selectbox(
                "",
                options=month_options,
                format_func=lambda x: "All" if x=="All" else datetime(2000, x, 1).strftime('%B'),
                index=default_index,
                label_visibility="collapsed",
                key=f"presumptive_month_select_{reset_count}"
            )
            st.session_state["report_filters_presumptive"]["selected_month_presumptive"] = new_month

        # Year
        with col_year:
            st.markdown('<div class="filter-label">Year</div>', unsafe_allow_html=True)
            current_year = st.session_state["report_filters_presumptive"].get(
                "selected_year_presumptive", datetime.today().year
            )
            year_range = list(range(2020, datetime.today().year+2))
            default_index = year_range.index(current_year) if current_year in year_range else len(year_range)-1
            new_year = st.selectbox(
                "",
                options=year_range,
                index=default_index,
                label_visibility="collapsed",
                key=f"presumptive_year_select_{reset_count}"
            )
            st.session_state["report_filters_presumptive"]["selected_year_presumptive"] = new_year

        # Reset 
        with col_reset:
            st.markdown('<div style="height:22px;"></div>', unsafe_allow_html=True)
            if st.button("Reset Filters", key="reset_presumptive_filters"):
                st.session_state["report_filters_presumptive"]["selected_month_presumptive"] = datetime.today().month
                st.session_state["report_filters_presumptive"]["selected_year_presumptive"] = datetime.today().year
                st.session_state["presumptive_reset_count"] += 1  # increment to force new keys
                st.rerun()

    selected_month_presumptive = st.session_state["report_filters_presumptive"]["selected_month_presumptive"]
    selected_year_presumptive = st.session_state["report_filters_presumptive"]["selected_year_presumptive"]

    # Fetch AI report data with filters
    ai_report_data = fetch_ai_report_data(selected_month_presumptive, selected_year_presumptive)

    # Show AI report metrics as styled cards
    st.markdown(f"""
        <div class="card-container">
            <div class="report-card">
                <div class="line-1">Total Flagged Patients</div>
                <div class="line-2">{ai_report_data['Total Flagged Patients']}</div>
                <div class="line-3">Presumptive Cases</div>
            </div>
            <div class="report-card">
                <div class="line-1">AI Accuracy Rate</div>
                <div class="line-2">{ai_report_data['AI Accuracy Rate']}</div>
                <div class="line-3">Diagnosis Match Rate</div>
            </div>
            <div class="report-card">
                <div class="line-1">Pending Confirmations</div>
                <div class="line-2">{ai_report_data['Pending Confirmations']}</div>
                <div class="line-3">Needs Review</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    # Generate PDF/Excel bytes for AI report (using enhanced functions)
    filter_info = format_filter_info(selected_month_presumptive, selected_year_presumptive)
    ai_pdf_bytes = generate_ai_pdf("AI Presumptive TB Report", ai_report_data, filter_info, 
                                    selected_month_presumptive, selected_year_presumptive)
    ai_excel_bytes = generate_ai_excel("AI Presumptive TB Report", ai_report_data, filter_info,
                                        selected_month_presumptive, selected_year_presumptive)

    st.markdown('<div style="height:25px;"></div>', unsafe_allow_html=True)

    # Export buttons
    col_empty, col_export_pdf, col_export_excel = st.columns([6, 1, 1])
    with col_export_pdf:
        st.download_button("Export PDF", data=ai_pdf_bytes, file_name=f"AI_Presumptive_TB_Report_{today_str}.pdf", mime="application/pdf")
    with col_export_excel:
        st.download_button("Export Excel", data=ai_excel_bytes, file_name=f"AI_Presumptive_TB_Report_{today_str}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    st.markdown('<div style="height:35px;"></div>', unsafe_allow_html=True)


    # --- Confirmed TB Cases Report Block ---

    # Display Confirmed TB Cases report title + reset button (under filters)
    col_title, col_filters = st.columns([6, 6])
    with col_title:
        st.markdown('<div class="card-title">Confirmed TB Cases Report</div>', unsafe_allow_html=True)

    with col_filters:
        col_month, col_year, col_reset = st.columns([1,1,0.6])

        # Initialize/reset counter
        if "confirmed_reset_count" not in st.session_state:
            st.session_state["confirmed_reset_count"] = 0
        reset_count = st.session_state["confirmed_reset_count"]

        # Month
        with col_month:
            st.markdown('<div class="filter-label">Month</div>', unsafe_allow_html=True)
            current_stored = st.session_state["report_filters_confirmed"].get(
                "selected_month_confirmed", datetime.today().month
            )
            month_options = ["All"] + list(range(1, 13))
            default_index = 0 if current_stored == "All" else month_options.index(current_stored)
            new_month = st.selectbox(
                "",
                options=month_options,
                format_func=lambda x: "All" if x=="All" else datetime(2000, x, 1).strftime('%B'),
                index=default_index,
                label_visibility="collapsed",
                key=f"confirmed_month_select_{reset_count}"
            )
            st.session_state["report_filters_confirmed"]["selected_month_confirmed"] = new_month

        # Year
        with col_year:
            st.markdown('<div class="filter-label">Year</div>', unsafe_allow_html=True)
            current_year = st.session_state["report_filters_confirmed"].get(
                "selected_year_confirmed", datetime.today().year
            )
            year_range = list(range(2020, datetime.today().year+2))
            default_index = year_range.index(current_year) if current_year in year_range else len(year_range)-1
            new_year = st.selectbox(
                "",
                options=year_range,
                index=default_index,
                label_visibility="collapsed",
                key=f"confirmed_year_select_{reset_count}"
            )
            st.session_state["report_filters_confirmed"]["selected_year_confirmed"] = new_year

        # Reset
        with col_reset:
            st.markdown('<div style="height:22px;"></div>', unsafe_allow_html=True)
            if st.button("Reset Filters", key="reset_confirmed_filters"):
                st.session_state["report_filters_confirmed"]["selected_month_confirmed"] = datetime.today().month
                st.session_state["report_filters_confirmed"]["selected_year_confirmed"] = datetime.today().year
                st.session_state["confirmed_reset_count"] += 1
                st.rerun()

    selected_month_confirmed = st.session_state["report_filters_confirmed"]["selected_month_confirmed"]
    selected_year_confirmed = st.session_state["report_filters_confirmed"]["selected_year_confirmed"]

    # Fetch confirmed TB cases data with filters
    confirmed_cases_data = fetch_confirmed_cases_data(selected_month_confirmed, selected_year_confirmed)

    # Show Confirmed cases as styled cards
    st.markdown(f"""
        <div class="card-container">
            <div class="report-card">
                <div class="line-1">Total Confirmed Cases</div>
                <div class="line-2">{confirmed_cases_data['Total Confirmed Cases'].split()[0]}</div>
                <div class="line-3">Confirmed TB Cases</div>
            </div>
            <div class="report-card">
                <div class="line-1">Positive Cases</div>
                <div class="line-2">{confirmed_cases_data['Positive Cases'].split()[0]}</div>
                <div class="line-3">{confirmed_cases_data['Positive Cases'].split(' ', 1)[1]}</div>
            </div>
            <div class="report-card">
                <div class="line-1">Negative Cases</div>
                <div class="line-2">{confirmed_cases_data['Negative Cases'].split()[0]}</div>
                <div class="line-3">{confirmed_cases_data['Negative Cases'].split(' ', 1)[1]}</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    # Generate PDF/Excel bytes for Confirmed TB Cases report (using enhanced functions)
    filter_info = format_filter_info(selected_month_confirmed, selected_year_confirmed)
    cases_pdf = generate_confirmed_pdf("Confirmed TB Cases Report", confirmed_cases_data, filter_info,
                                        selected_month_confirmed, selected_year_confirmed)
    cases_excel_bytes = generate_confirmed_excel("Confirmed TB Cases Report", confirmed_cases_data, filter_info,
                                                  selected_month_confirmed, selected_year_confirmed)

    st.markdown('<div style="height:25px;"></div>', unsafe_allow_html=True)

    # Export buttons at bottom right
    col_empty, col_export_pdf, col_export_excel = st.columns([6, 1, 1])
    with col_export_pdf:
        st.download_button("Export PDF", data=cases_pdf, file_name=f"Confirmed_TB_Cases_Report_{today_str}.pdf", mime="application/pdf")
    with col_export_excel:
        st.download_button("Export Excel", data=cases_excel_bytes, file_name=f"Confirmed_TB_Cases_Report_{today_str}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")